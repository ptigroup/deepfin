"""Tests for Excel export functionality."""

from decimal import Decimal
from io import BytesIO
from pathlib import Path
from tempfile import NamedTemporaryFile
from unittest.mock import MagicMock

import pytest
from openpyxl import load_workbook

from app.consolidation.exporter import ConsolidationExporter
from app.consolidation.models import ConsolidatedStatement, PeriodComparison


class TestConsolidationExporter:
    """Test Excel export functionality."""

    @pytest.fixture
    def exporter(self):
        """Create exporter instance."""
        return ConsolidationExporter()

    @pytest.fixture
    def mock_statement(self):
        """Create mock consolidated statement."""
        stmt = MagicMock(spec=ConsolidatedStatement)
        stmt.id = 1
        stmt.name = "Q1-Q4 2024 Consolidated"
        stmt.company_name = "Acme Corp"
        stmt.start_period = "2024-01-01"
        stmt.end_period = "2024-12-31"
        stmt.currency = "USD"
        stmt.period_count = 4
        stmt.total_line_items = 120
        stmt.description = "Annual consolidation"
        return stmt

    @pytest.fixture
    def mock_comparisons(self):
        """Create mock period comparisons."""
        comp1 = MagicMock(spec=PeriodComparison)
        comp1.line_item_name = "Total Revenue"
        comp1.current_period = "Q4 2024"
        comp1.previous_period = "Q3 2024"
        comp1.current_value = Decimal("500000.00")
        comp1.previous_value = Decimal("450000.00")
        comp1.change_amount = Decimal("50000.00")
        comp1.change_percentage = Decimal("11.11")
        comp1.is_favorable = True

        comp2 = MagicMock(spec=PeriodComparison)
        comp2.line_item_name = "Operating Expenses"
        comp2.current_period = "Q4 2024"
        comp2.previous_period = "Q3 2024"
        comp2.current_value = Decimal("200000.00")
        comp2.previous_value = Decimal("180000.00")
        comp2.change_amount = Decimal("20000.00")
        comp2.change_percentage = Decimal("11.11")
        comp2.is_favorable = False

        return [comp1, comp2]

    def test_export_to_bytesio(self, exporter, mock_statement, mock_comparisons):
        """Test exporting to BytesIO buffer."""
        result = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        assert isinstance(result, BytesIO)
        assert result.tell() == 0
        content = result.read()
        assert len(content) > 0
        assert content[:2] == b"PK"  # ZIP signature

    def test_export_to_file(self, exporter, mock_statement, mock_comparisons):
        """Test exporting to file path."""
        with NamedTemporaryFile(delete=False, suffix=".xlsx") as tmp_file:
            tmp_path = Path(tmp_file.name)

        try:
            result = exporter.export_consolidated_statement(
                mock_statement,
                mock_comparisons,
                output_path=tmp_path,
            )

            assert result == tmp_path
            assert tmp_path.exists()
            assert tmp_path.stat().st_size > 0
        finally:
            if tmp_path.exists():
                tmp_path.unlink()

    def test_export_workbook_structure(self, exporter, mock_statement, mock_comparisons):
        """Test exported workbook structure."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        # Verify sheet name
        assert ws.title == "Consolidation"

        # Verify header section
        assert ws["A1"].value == "Q1-Q4 2024 Consolidated"
        assert ws["A3"].value == "Company:"
        assert ws["B3"].value == "Acme Corp"
        assert ws["A5"].value == "Currency:"
        assert ws["B5"].value == "USD"

        wb.close()

    def test_export_comparison_table_headers(self, exporter, mock_statement, mock_comparisons):
        """Test comparison table headers."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        # Verify table headers (row 10)
        expected_headers = [
            "Line Item",
            "Current Period",
            "Current Value",
            "Previous Period",
            "Previous Value",
            "Change Amount",
            "Change %",
            "Favorable",
        ]

        for col_idx, expected_header in enumerate(expected_headers, start=1):
            cell_value = ws.cell(row=10, column=col_idx).value
            assert cell_value == expected_header

        wb.close()

    def test_export_comparison_data(self, exporter, mock_statement, mock_comparisons):
        """Test comparison data in exported file."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        # Verify first comparison (row 11)
        assert ws.cell(row=11, column=1).value == "Total Revenue"
        assert ws.cell(row=11, column=8).value == "Yes"

        # Verify second comparison (row 12)
        assert ws.cell(row=12, column=1).value == "Operating Expenses"
        assert ws.cell(row=12, column=8).value == "No"

        wb.close()

    def test_export_no_comparisons(self, exporter, mock_statement):
        """Test exporting statement with no comparisons."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            [],
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        assert ws["A1"].value == "Q1-Q4 2024 Consolidated"
        assert ws["A10"].value == "No period comparisons available"

        wb.close()

    def test_export_number_formatting(self, exporter, mock_statement, mock_comparisons):
        """Test number formatting in exported file."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        # Check currency formatting
        assert ws.cell(row=11, column=3).number_format == "#,##0.00"

        # Check percentage formatting
        assert ws.cell(row=11, column=7).number_format == "0.00%"

        wb.close()

    def test_export_cell_styling(self, exporter, mock_statement, mock_comparisons):
        """Test cell styling in exported file."""
        buffer = exporter.export_consolidated_statement(
            mock_statement,
            mock_comparisons,
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        # Check header styling
        header_cell = ws.cell(row=10, column=1)
        assert header_cell.font.bold is True

        # Check favorable color coding (green-ish)
        favorable_cell = ws.cell(row=11, column=8)
        color_rgb = favorable_cell.fill.start_color.rgb
        # Should be green (starts with FFC6)
        assert color_rgb.startswith("FFC6") or color_rgb.startswith("00C6")

        # Check unfavorable color coding (red-ish)
        unfavorable_cell = ws.cell(row=12, column=8)
        color_rgb = unfavorable_cell.fill.start_color.rgb
        # Should be red (contains FFC7 or FF00)
        assert "FFC7" in color_rgb or "FF00" in color_rgb or color_rgb.startswith("FFFF")

        wb.close()

    def test_export_favorable_null(self, exporter, mock_statement):
        """Test exporting comparison with null favorable indicator."""
        comp = MagicMock(spec=PeriodComparison)
        comp.line_item_name = "Total Assets"
        comp.current_period = "Q4 2024"
        comp.previous_period = "Q3 2024"
        comp.current_value = Decimal("2000000.00")
        comp.previous_value = Decimal("1800000.00")
        comp.change_amount = Decimal("200000.00")
        comp.change_percentage = Decimal("11.11")
        comp.is_favorable = None

        buffer = exporter.export_consolidated_statement(
            mock_statement,
            [comp],
            output_path=None,
        )

        wb = load_workbook(buffer)
        ws = wb.active

        assert ws.cell(row=11, column=8).value == "N/A"

        wb.close()
