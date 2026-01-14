"""Excel export functionality for consolidated financial data."""

from io import BytesIO
from pathlib import Path

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter

from app.consolidation.models import ConsolidatedStatement, PeriodComparison
from app.core.logging import get_logger

logger = get_logger(__name__)


class ExcelExportError(Exception):
    """Exception raised during Excel export."""

    pass


class ConsolidationExporter:
    """Exports consolidated financial data to formatted Excel files.

    Handles:
    - Creating formatted Excel workbooks
    - Styling headers, data rows, and totals
    - Exporting consolidated statements with comparisons
    - Number formatting for financial data
    """

    def __init__(self):
        """Initialize Excel exporter."""
        # Define styles
        self.header_font = Font(bold=True, size=12, color="FFFFFF")
        self.header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        self.border = Border(
            left=Side(style="thin"),
            right=Side(style="thin"),
            top=Side(style="thin"),
            bottom=Side(style="thin"),
        )
        self.center_alignment = Alignment(horizontal="center", vertical="center")
        self.right_alignment = Alignment(horizontal="right", vertical="center")

    def export_consolidated_statement(
        self,
        statement: ConsolidatedStatement,
        comparisons: list[PeriodComparison],
        output_path: Path | None = None,
    ) -> BytesIO | Path:
        """Export consolidated statement with comparisons to Excel.

        Args:
            statement: Consolidated statement to export
            comparisons: List of period comparisons
            output_path: Optional output file path (if None, returns BytesIO)

        Returns:
            BytesIO buffer or Path to saved file

        Raises:
            ExcelExportError: If export fails
        """
        try:
            # Create workbook
            wb = Workbook()
            ws = wb.active
            ws.title = "Consolidation"

            # Write header section
            self._write_header_section(ws, statement)

            # Write comparison data
            if comparisons:
                self._write_comparison_section(ws, comparisons)
            else:
                ws["A10"] = "No period comparisons available"

            # Auto-size columns
            self._auto_size_columns(ws)

            # Save or return buffer
            if output_path:
                wb.save(output_path)
                logger.info(
                    "Exported consolidated statement to file",
                    extra={"statement_id": statement.id, "path": str(output_path)},
                )
                return output_path
            else:
                buffer = BytesIO()
                wb.save(buffer)
                buffer.seek(0)
                logger.info(
                    "Exported consolidated statement to buffer",
                    extra={"statement_id": statement.id},
                )
                return buffer

        except Exception as e:
            logger.error(
                "Failed to export consolidated statement",
                extra={"statement_id": statement.id, "error": str(e)},
                exc_info=True,
            )
            raise ExcelExportError(f"Failed to export: {str(e)}") from e

    def _write_header_section(self, ws, statement: ConsolidatedStatement) -> None:
        """Write header section with consolidation metadata.

        Args:
            ws: Worksheet to write to
            statement: Consolidated statement
        """
        # Title
        ws["A1"] = statement.name
        ws["A1"].font = Font(bold=True, size=14)

        # Metadata
        ws["A3"] = "Company:"
        ws["B3"] = statement.company_name
        ws["A3"].font = Font(bold=True)

        ws["A4"] = "Period Range:"
        ws["B4"] = f"{statement.start_period} to {statement.end_period}"
        ws["A4"].font = Font(bold=True)

        ws["A5"] = "Currency:"
        ws["B5"] = statement.currency
        ws["A5"].font = Font(bold=True)

        ws["A6"] = "Total Periods:"
        ws["B6"] = statement.period_count
        ws["A6"].font = Font(bold=True)

        if statement.description:
            ws["A7"] = "Description:"
            ws["B7"] = statement.description
            ws["A7"].font = Font(bold=True)

    def _write_comparison_section(
        self,
        ws,
        comparisons: list[PeriodComparison],
    ) -> None:
        """Write period comparison data section.

        Args:
            ws: Worksheet to write to
            comparisons: List of period comparisons
        """
        # Start row for comparison table
        start_row = 10

        # Write table headers
        headers = [
            "Line Item",
            "Current Period",
            "Current Value",
            "Previous Period",
            "Previous Value",
            "Change Amount",
            "Change %",
            "Favorable",
        ]

        for col_idx, header in enumerate(headers, start=1):
            cell = ws.cell(row=start_row, column=col_idx, value=header)
            cell.font = self.header_font
            cell.fill = self.header_fill
            cell.border = self.border
            cell.alignment = self.center_alignment

        # Write comparison data
        for row_idx, comparison in enumerate(comparisons, start=start_row + 1):
            ws.cell(row=row_idx, column=1, value=comparison.line_item_name)
            ws.cell(row=row_idx, column=2, value=comparison.current_period)
            ws.cell(row=row_idx, column=3, value=float(comparison.current_value))
            ws.cell(row=row_idx, column=4, value=comparison.previous_period)
            ws.cell(row=row_idx, column=5, value=float(comparison.previous_value))
            ws.cell(row=row_idx, column=6, value=float(comparison.change_amount))
            ws.cell(row=row_idx, column=7, value=float(comparison.change_percentage))
            ws.cell(
                row=row_idx,
                column=8,
                value="Yes"
                if comparison.is_favorable
                else "No"
                if comparison.is_favorable is False
                else "N/A",
            )

            # Apply formatting
            for col in range(1, 9):
                cell = ws.cell(row=row_idx, column=col)
                cell.border = self.border

                # Right-align numbers
                if col in {3, 5, 6, 7}:
                    cell.alignment = self.right_alignment
                    # Number format
                    if col in {3, 5, 6}:
                        cell.number_format = "#,##0.00"
                    elif col == 7:
                        cell.number_format = "0.00%"

                # Highlight favorable changes
                if col == 8 and comparison.is_favorable:
                    cell.fill = PatternFill(
                        start_color="C6EFCE",
                        end_color="C6EFCE",
                        fill_type="solid",
                    )
                elif col == 8 and comparison.is_favorable is False:
                    cell.fill = PatternFill(
                        start_color="FFC7CE",
                        end_color="FFC7CE",
                        fill_type="solid",
                    )

    def _auto_size_columns(self, ws) -> None:
        """Auto-size columns based on content.

        Args:
            ws: Worksheet to auto-size
        """
        for column_cells in ws.columns:
            max_length = 0
            column_letter = get_column_letter(column_cells[0].column)

            for cell in column_cells:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))

            adjusted_width = min(max_length + 2, 50)  # Cap at 50
            ws.column_dimensions[column_letter].width = adjusted_width
