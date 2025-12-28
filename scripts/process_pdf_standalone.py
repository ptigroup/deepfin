#!/usr/bin/env python3
"""Standalone PDF processor that works WITHOUT database.

This script directly calls LLMWhisperer and extraction logic without
database persistence - perfect for demos and validation.

Usage:
    python scripts/process_pdf_standalone.py path/to/statement.pdf
"""

import argparse
import json
import sys
from datetime import datetime
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Load environment variables from .env file
from dotenv import load_dotenv
load_dotenv()

from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from unstract.llmwhisperer import LLMWhispererClientV2
from unstract.llmwhisperer.client_v2 import LLMWhispererClientException

from app.core.logging import get_logger
from app.extraction.pydantic_extractor import PydanticExtractor
from app.extraction.page_detector import PageDetector, FinancialStatementType

logger = get_logger(__name__)


def process_pdf_standalone(
    pdf_path: Path,
    output_dir: Path,
) -> dict:
    """Process PDF without database - direct extraction only.

    Args:
        pdf_path: Path to PDF file
        output_dir: Directory for output files

    Returns:
        Processing result with metadata
    """
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    print(f"\n{'='*70}")
    print(f"Processing: {pdf_path.name}")
    print(f"{'='*70}\n")

    start_time = datetime.now()

    try:
        # Step 0: Detect pages for large PDFs (cost optimization)
        import fitz  # PyMuPDF

        doc = fitz.open(str(pdf_path))
        total_pages = len(doc)
        doc.close()

        pages_to_extract = None
        detected_statement_type = None

        if total_pages > 10:  # Large PDF - use page detection
            print(f">> Step 0: Detecting table pages (PDF has {total_pages} pages)...")
            detector = PageDetector()

            # Use new method which returns Dict[FinancialStatementType, List[int]]
            detected_pages = detector.detect_financial_tables(pdf_path)

            if detected_pages and FinancialStatementType.INCOME_STATEMENT in detected_pages:
                # Get primary page(s) for income statement
                primary_pages = detected_pages[FinancialStatementType.INCOME_STATEMENT]

                # Expand with surrounding pages for context
                page_ranges = detector.get_page_ranges_for_extraction(detected_pages)
                pages_to_extract = page_ranges.get(FinancialStatementType.INCOME_STATEMENT, primary_pages)
                detected_statement_type = "income_statement"

                print(f"OK: Detected income statement on page(s) {primary_pages}")
                print(f"   With context: Extracting {len(pages_to_extract)} pages instead of {total_pages} pages")
                print(f"   Cost optimization: {100 - (len(pages_to_extract)/total_pages)*100:.1f}% reduction")
            else:
                print("WARNING: No income statement detected, extracting all pages")
        else:
            print(f">> Step 0: Small PDF ({total_pages} pages) - extracting all pages")

        # Step 1: Extract text with LLMWhisperer V2
        print("\n>> Step 1: Extracting text with LLMWhisperer...")
        client = LLMWhispererClientV2()

        whisper_params = {
            "file_path": str(pdf_path),
            "mode": "form",
            "output_mode": "layout_preserving",
            "wait_for_completion": True,
            "wait_timeout": 200,
        }

        # Add pages_to_extract if we detected specific pages
        if pages_to_extract:
            whisper_params["pages_to_extract"] = ",".join(map(str, pages_to_extract))
            print(f"   Extracting only pages: {whisper_params['pages_to_extract']}")

        result = client.whisper(**whisper_params)

        # Debug: Print response structure
        logger.debug("Whisper result keys", extra={"keys": list(result.keys())})
        if "extraction" in result:
            logger.debug("Extraction keys", extra={"keys": list(result["extraction"].keys())})

        # Extract text from V2 response format
        # V2 format: result["extraction"]["result_text"] contains the extracted text
        extraction = result.get("extraction", {})
        raw_text = extraction.get("result_text", "")

        print(f"OK: Extracted {len(raw_text)} characters of raw text")

        # Debug: Show first 200 chars if available
        if raw_text:
            logger.debug("Raw text preview", extra={"preview": raw_text[:200]})

            # Save raw text for inspection
            raw_text_path = output_dir / "raw_text" / f"{pdf_path.stem}_raw.txt"
            raw_text_path.parent.mkdir(parents=True, exist_ok=True)
            with open(raw_text_path, "w", encoding="utf-8") as f:
                f.write(raw_text)
            print(f"DEBUG: Saved raw text to {raw_text_path}")

        # Step 2: Extract structured data with Pydantic AI
        print("\n>> Step 2: Extracting structured data with Pydantic AI...")
        extractor = PydanticExtractor()

        # Extract structured data
        parsed_data = extractor.extract_from_text(raw_text)

        # Extract details from parsed data
        statement_type = parsed_data.get("statement_type", "financial_statement")
        company_name = parsed_data.get("company_name", pdf_path.stem)
        fiscal_year = parsed_data.get("fiscal_year")
        periods = parsed_data.get("periods", [])
        line_items = parsed_data.get("line_items", [])

        print(f"OK: Extracted structured data")
        print(f"  - Type: {statement_type}")
        print(f"  - Company: {company_name}")
        print(f"  - Periods: {len(periods)}")
        print(f"  - Line items: {len(line_items)}")

        # Prepare metadata
        end_time = datetime.now()
        processing_time = (end_time - start_time).total_seconds()

        metadata = {
            "pdf_name": pdf_path.name,
            "pdf_size_kb": round(pdf_path.stat().st_size / 1024, 2),
            "pdf_total_pages": total_pages,
            "document_type": statement_type,
            "company_name": company_name,
            "fiscal_year": fiscal_year,
            "extracted_at": end_time.isoformat(),
            "processing_time_seconds": round(processing_time, 2),
            "periods_count": len(periods),
            "line_items_count": len(line_items),
            "extraction_mode": "page_detection" if pages_to_extract else "full_pdf",
            "pages_extracted": pages_to_extract if pages_to_extract else list(range(1, total_pages + 1)),
            "characters_extracted": len(raw_text),
        }

        # Prepare output data
        output_data = {
            "metadata": metadata,
            "data": {
                "company_name": company_name,
                "statement_type": statement_type,
                "fiscal_year": fiscal_year,
                "currency": parsed_data.get("currency", "USD"),
                "periods": periods,
                "line_items": line_items,
            }
        }

        # Step 3: Save JSON output
        print("\n>> Step 3: Saving outputs...")
        json_dir = output_dir / "json"
        json_dir.mkdir(parents=True, exist_ok=True)
        json_path = json_dir / f"{pdf_path.stem}.json"

        with open(json_path, "w", encoding="utf-8") as f:
            json.dump(output_data, f, indent=2, ensure_ascii=False)

        print(f"OK: Saved JSON: {json_path}")

        # Step 4: Save Excel output
        excel_dir = output_dir / "excel"
        excel_dir.mkdir(parents=True, exist_ok=True)
        excel_path = excel_dir / f"{pdf_path.stem}.xlsx"

        # Create Excel workbook
        wb = Workbook()
        ws = wb.active
        ws.title = statement_type[:31]  # Excel max 31 chars

        # Styling
        header_font = Font(bold=True, size=12)
        center_align = Alignment(horizontal="center")
        right_align = Alignment(horizontal="right")

        # Write header
        ws.cell(1, 1, "Account").font = header_font
        for idx, period in enumerate(periods, start=2):
            ws.cell(1, idx, period).font = header_font
            ws.cell(1, idx).alignment = center_align

        # Write line items
        for row_idx, item in enumerate(line_items, start=2):
            # Apply indentation
            indent_level = item.get("indent_level", 0)
            account_name = item.get("account_name", "")
            values = item.get("values", [])

            # Note: Pydantic AI preserves indentation in account_name
            # so we don't need to add extra indent here
            ws.cell(row_idx, 1, account_name)

            # Write values
            for col_idx, value in enumerate(values, start=2):
                cell = ws.cell(row_idx, col_idx, value)
                cell.alignment = right_align

        # Auto-adjust column widths
        for col in ws.columns:
            max_length = 0
            column = col[0].column_letter
            for cell in col:
                try:
                    cell_value = str(cell.value) if cell.value else ""
                    if len(cell_value) > max_length:
                        max_length = len(cell_value)
                except:
                    pass
            adjusted_width = min(max_length + 2, 50)
            ws.column_dimensions[column].width = adjusted_width

        # Save workbook
        wb.save(excel_path)
        print(f"OK: Saved Excel: {excel_path}")

        # Save metadata
        metadata_dir = output_dir / "metadata"
        metadata_dir.mkdir(parents=True, exist_ok=True)
        metadata_path = metadata_dir / f"{pdf_path.stem}_metadata.json"

        with open(metadata_path, "w", encoding="utf-8") as f:
            json.dump(metadata, f, indent=2)

        print(f"OK: Saved metadata: {metadata_path}")

        print(f"\n{'='*70}")
        print(f"SUCCESS - Completed in {processing_time:.2f} seconds")
        print(f"{'='*70}\n")

        return {
            "status": "success",
            "pdf": pdf_path.name,
            "metadata": metadata,
            "outputs": {
                "json": str(json_path),
                "excel": str(excel_path),
                "metadata": str(metadata_path),
            }
        }

    except Exception as e:
        print(f"\nERROR: {str(e)}")
        logger.error(
            "PDF processing failed",
            extra={"pdf": pdf_path.name, "error": str(e)},
            exc_info=True
        )

        return {
            "status": "error",
            "pdf": pdf_path.name,
            "error": str(e),
        }


def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Process financial PDF (standalone - no database required)"
    )
    parser.add_argument(
        "pdf_path",
        type=Path,
        help="Path to PDF file"
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("samples/output"),
        help="Output directory (default: samples/output)"
    )

    args = parser.parse_args()

    try:
        result = process_pdf_standalone(
            pdf_path=args.pdf_path,
            output_dir=args.output,
        )

        # Exit with appropriate code
        sys.exit(0 if result["status"] == "success" else 1)

    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
        sys.exit(130)
    except Exception as e:
        print(f"\n\nFatal error: {str(e)}")
        logger.error("Fatal error in main", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    main()
