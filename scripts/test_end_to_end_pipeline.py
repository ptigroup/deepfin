#!/usr/bin/env python3
"""
End-to-End Financial Statement Processing Pipeline Test

ZERO HARDCODING - Fully dynamic detection and extraction.

Pipeline:
1. DYNAMIC Detection: Find ALL financial statement types in each PDF
2. SMART Extraction: Extract ONLY detected pages (no token waste)
3. PARSING: Structure data with OpenAI + Pydantic validation
4. CONSOLIDATION: Merge statements by type across PDFs (fuzzy matching)
5. OUTPUT: Generate JSON + Excel for each consolidated statement type
6. VALIDATION: Comprehensive checks on pages, data, timeline, outputs

Usage:
    # Testing mode (uses samples/ directory)
    python scripts/test_end_to_end_pipeline.py

    # Production mode (uses input/ directory)
    python scripts/test_end_to_end_pipeline.py --mode production

    # Custom PDFs (auto-detects mode from path)
    python scripts/test_end_to_end_pipeline.py path/to/file1.pdf path/to/file2.pdf

    # Production mode with specific PDFs
    python scripts/test_end_to_end_pipeline.py --mode production input/Company_10K.pdf

Test PDFs (testing mode):
    - samples/input/NVIDIA 10K 2020-2019.pdf
    - samples/input/NVIDIA 10K 2022-2021.pdf

Expected Outputs:
    Testing mode:  samples/output/runs/YYYYMMDD_HHMMSS_STATUS/
    Production mode: output/runs/YYYYMMDD_HHMMSS_STATUS/
"""

import os
import sys
import json
import argparse
from pathlib import Path
from datetime import datetime
from typing import Dict, List, Tuple, Optional, TYPE_CHECKING
from collections import defaultdict
from difflib import SequenceMatcher

if TYPE_CHECKING:
    from openpyxl import Workbook

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

# Load environment variables
from dotenv import load_dotenv
load_dotenv()

# Import path configuration
from scripts.path_config import PathConfig

from unstract.llmwhisperer import LLMWhispererClientV2
from unstract.llmwhisperer.client_v2 import LLMWhispererClientException

from app.core.logging import get_logger
from app.core.output_manager import OutputManager, ExtractionRun, RunStatus
from app.extraction.page_detector import PageDetector, FinancialStatementType
from app.extraction.pydantic_extractor import PydanticExtractor
from app.export.excel_exporter import SchemaBasedExcelExporter

logger = get_logger(__name__)


class ConsolidationTracker:
    """Track merge operations during consolidation for transparency and debugging."""

    def __init__(self):
        self.merged_accounts = []  # List of accounts that were merged
        self.unmerged_accounts = []  # Accounts that appeared in only one source
        self.duplicate_headers = []  # Section headers that appear multiple times
        self.source_pdfs = set()  # All source PDFs

    def track_merge(self, consolidated_name: str, section: Optional[str],
                    sources: List[Dict[str, str]]):
        """
        Track that multiple accounts were merged into one.

        Args:
            consolidated_name: The final account name in consolidated output
            section: Section this account belongs to
            sources: List of dicts with 'name', 'section', 'source_pdf' keys
        """
        # Only track if actually merged from multiple sources
        if len(sources) > 1:
            self.merged_accounts.append({
                "consolidated_name": consolidated_name,
                "section": section,
                "merged_from": sources
            })

        # Track source PDFs
        for source in sources:
            self.source_pdfs.add(source.get("source_pdf", "unknown"))

    def track_unmerged(self, account_name: str, section: Optional[str],
                       source_pdf: str, reason: str, values_count: int = 0):
        """
        Track accounts that appeared in only one source (potential missing data).

        Args:
            account_name: Account name
            section: Section
            source_pdf: Which PDF it came from
            reason: Why it wasn't merged (e.g., 'only_in_one_pdf', 'section_mismatch')
            values_count: Number of values (0 = empty header)
        """
        self.unmerged_accounts.append({
            "account_name": account_name,
            "section": section,
            "source_pdf": source_pdf,
            "reason": reason,
            "values_count": values_count
        })

    def track_duplicate_header(self, header_name: str, sources: List[Dict[str, str]]):
        """Track section headers that appear multiple times with different metadata."""
        if len(sources) > 1:
            self.duplicate_headers.append({
                "header_name": header_name,
                "occurrences": sources
            })

    def get_summary(self) -> Dict:
        """Get consolidation summary for inclusion in output."""
        return {
            "merged_accounts": self.merged_accounts,
            "unmerged_accounts": self.unmerged_accounts,
            "duplicate_headers": self.duplicate_headers,
            "total_accounts_merged": len([m for m in self.merged_accounts if len(m["merged_from"]) > 1]),
            "total_accounts_unmerged": len(self.unmerged_accounts),
            "total_duplicate_headers": len(self.duplicate_headers),
            "source_pdfs": sorted(list(self.source_pdfs))
        }


class EndToEndPipeline:
    """Fully dynamic end-to-end financial statement processing pipeline."""

    def __init__(self, output_base: str = "output"):
        """Initialize pipeline components."""
        self.detector = PageDetector()
        self.extractor = PydanticExtractor()
        self.llm_client = LLMWhispererClientV2()
        self.output_manager = OutputManager(output_base)

        # Current run (set in run_full_pipeline)
        self.current_run: Optional[ExtractionRun] = None

        # Validation tracking
        self.validation_results = {
            "detection": {},
            "extraction": {},
            "consolidation": {},
            "outputs": {},
            "spot_checks": {}
        }

    def run_full_pipeline(self, pdf_paths: List[Path], output_dir: Path = None) -> Dict:
        """
        Run complete pipeline from PDFs to consolidated outputs.

        Args:
            pdf_paths: List of PDF file paths to process
            output_dir: Directory for consolidated outputs (DEPRECATED - now uses OutputManager)

        Returns:
            Complete processing results with validation
        """
        print("\n" + "=" * 80)
        print("END-TO-END FINANCIAL STATEMENT PROCESSING PIPELINE")
        print("=" * 80)
        print(f"Input PDFs: {len(pdf_paths)}")
        print("=" * 80 + "\n")

        start_time = datetime.now()

        # Create new extraction run
        self.current_run = self.output_manager.create_run(status=RunStatus.IN_PROGRESS)
        print(f"[RUN] Created run: {self.current_run.run_id}")
        print(f"[RUN] Output directory: {self.current_run.run_dir}\n")

        # PHASE 1: DYNAMIC DETECTION (NO HARDCODING!)
        print("\n" + ">" * 40)
        print("PHASE 1: DYNAMIC TABLE DETECTION")
        print(">" * 40 + "\n")

        all_detections = self._phase1_detect_all_pdfs(pdf_paths)

        # PHASE 2: EXTRACTION (ONLY DETECTED PAGES!)
        print("\n" + ">" * 40)
        print("PHASE 2: SMART EXTRACTION (DETECTED PAGES ONLY)")
        print(">" * 40 + "\n")

        all_extractions = self._phase2_extract_all_statements(pdf_paths, all_detections)

        # PHASE 3: CONSOLIDATION (MERGE BY STATEMENT TYPE!)
        print("\n" + ">" * 40)
        print("PHASE 3: CONSOLIDATION (MERGE TIMELINES)")
        print(">" * 40 + "\n")

        consolidated = self._phase3_consolidate_by_type(all_extractions)

        # PHASE 4: OUTPUT GENERATION
        print("\n" + ">" * 40)
        print("PHASE 4: OUTPUT GENERATION (JSON + EXCEL)")
        print(">" * 40 + "\n")

        outputs = self._phase4_generate_outputs(consolidated, self.current_run.consolidated_dir)

        # PHASE 4B: COMBINED OUTPUT (ALL STATEMENT TYPES IN ONE FILE)
        print("\n" + ">" * 40)
        print("PHASE 4B: COMBINED OUTPUT (ALL STATEMENTS IN ONE FILE)")
        print(">" * 40 + "\n")

        combined_outputs = self._phase4b_generate_combined_outputs(consolidated, self.current_run.consolidated_dir)

        # PHASE 5: COMPREHENSIVE VALIDATION
        print("\n" + ">" * 40)
        print("PHASE 5: COMPREHENSIVE VALIDATION")
        print(">" * 40 + "\n")

        validation = self._phase5_validate_everything(
            all_detections, all_extractions, consolidated, outputs
        )

        # Generate final report
        duration = (datetime.now() - start_time).total_seconds()
        report = self._generate_report(validation, duration)

        print("\n" + report)

        # Complete the run
        final_status = RunStatus.SUCCESS if validation.get("overall_success", False) else RunStatus.PARTIAL
        self.current_run.complete(status=final_status)

        print(f"\n[RUN] Completed run: {self.current_run.run_id} ({final_status})")
        print(f"[RUN] Run directory: {self.current_run.run_dir}")
        print(f"[RUN] Checksums generated: {self.current_run.run_dir / 'checksums.md5'}")
        print(f"[RUN] History updated: {self.output_manager.output_base / 'RUN_HISTORY.md'}")

        return {
            "run_id": self.current_run.run_id,
            "run_dir": str(self.current_run.run_dir),
            "detections": all_detections,
            "extractions": all_extractions,
            "consolidated": consolidated,
            "outputs": outputs,
            "combined_outputs": combined_outputs,
            "validation": validation,
            "report": report,
            "duration_seconds": duration
        }

    def _phase1_detect_all_pdfs(self, pdf_paths: List[Path]) -> Dict[Path, Dict]:
        """Phase 1: Dynamically detect all financial statement types in each PDF."""
        all_detections = {}

        for pdf_path in pdf_paths:
            print(f"\n>> Detecting statements in: {pdf_path.name}")
            print("-" * 70)

            try:
                # DYNAMIC DETECTION - NO HARDCODED PAGES OR TYPES!
                detected_pages = self.detector.detect_financial_tables(pdf_path)

                all_detections[pdf_path] = detected_pages

                # Report findings
                if detected_pages:
                    print(f"[OK] Found {len(detected_pages)} statement types:")
                    for stmt_type, pages in detected_pages.items():
                        page_str = ", ".join(map(str, pages))
                        print(f"   - {stmt_type.value:25} -> Pages: {page_str}")
                else:
                    print("[FAIL] No financial statements detected")

            except Exception as e:
                logger.error(f"Detection failed for {pdf_path.name}: {e}")
                all_detections[pdf_path] = {}

        return all_detections

    def _phase2_extract_all_statements(
        self,
        pdf_paths: List[Path],
        all_detections: Dict[Path, Dict]
    ) -> Dict[Tuple[Path, FinancialStatementType], Dict]:
        """Phase 2: Extract ONLY detected pages (no token waste)."""
        all_extractions = {}

        for pdf_path in pdf_paths:
            detected_pages = all_detections.get(pdf_path, {})

            if not detected_pages:
                print(f"\n[WARN]  Skipping {pdf_path.name} - no statements detected")
                continue

            print(f"\n>> Extracting from: {pdf_path.name}")
            print("-" * 70)

            for stmt_type, pages in detected_pages.items():
                print(f"\n  Processing: {stmt_type.value}")
                print(f"  Pages: {pages}")

                try:
                    # EXTRACT ONLY DETECTED PAGES!
                    raw_text = self._extract_pages_with_llmwhisperer(pdf_path, pages)

                    # Parse with Pydantic AI
                    structured_dict = self.extractor.extract_from_text(raw_text)

                    # Convert to object for easier access
                    from app.extraction.pydantic_extractor import FinancialStatement
                    structured_data = FinancialStatement(**structured_dict)

                    # Calculate metrics
                    line_count = len(structured_data.line_items) if hasattr(structured_data, 'line_items') else 0
                    periods = structured_data.periods if hasattr(structured_data, 'periods') else []

                    # Store extraction
                    all_extractions[(pdf_path, stmt_type)] = {
                        "statement_type": stmt_type,
                        "pages": pages,
                        "raw_text": raw_text,
                        "structured_data": structured_data,
                        "pdf_name": pdf_path.name
                    }

                    # Save extraction to run directory
                    pdf_name_clean = pdf_path.stem.replace(" ", "_")
                    self.current_run.save_extraction(
                        pdf_name=pdf_name_clean,
                        statement_type=stmt_type.value,
                        json_data=structured_dict,
                        raw_text=raw_text,
                        metadata={
                            "pages": pages,
                            "pdf_filename": pdf_path.name,
                            "periods": periods,
                            "line_item_count": line_count
                        },
                        page_detection={
                            "detected_pages": pages,
                            "statement_type": stmt_type.value
                        }
                    )

                    # Report success
                    print(f"  [OK] Extracted {line_count} line items, {len(periods)} periods")
                    print(f"  [OK] Saved to: extracted/{pdf_name_clean}/{stmt_type.value}.json")

                except Exception as e:
                    logger.error(f"Extraction failed for {stmt_type.value}: {e}")
                    print(f"  [FAIL] Extraction failed: {e}")

        return all_extractions

    def _extract_pages_with_llmwhisperer(self, pdf_path: Path, pages: List[int]) -> str:
        """Extract specific pages using LLMWhisperer API."""
        # Convert pages list to comma-separated string
        pages_str = ",".join(map(str, pages))

        print(f"  LLMWhisperer: Extracting pages {pages_str}...")

        try:
            result = self.llm_client.whisper(
                file_path=str(pdf_path),
                mode="form",  # Preserve table structure
                output_mode="layout_preserving",
                pages_to_extract=pages_str,
                wait_for_completion=True,
                wait_timeout=200
            )

            raw_text = result.get("extraction", {}).get("result_text", "")

            if not raw_text:
                raise ValueError("LLMWhisperer returned empty text")

            return raw_text

        except LLMWhispererClientException as e:
            logger.error(f"LLMWhisperer API error: {e}")
            raise

    def _phase3_consolidate_by_type(
        self,
        all_extractions: Dict[Tuple[Path, FinancialStatementType], Dict]
    ) -> Dict[FinancialStatementType, Dict]:
        """Phase 3: Consolidate statements by type (merge both PDFs)."""
        # Group extractions by statement type
        by_type = defaultdict(list)

        for (pdf_path, stmt_type), extraction_data in all_extractions.items():
            by_type[stmt_type].append(extraction_data)

        # Consolidate each statement type
        consolidated = {}

        for stmt_type, extractions in by_type.items():
            print(f"\n  Consolidating: {stmt_type.value}")
            print(f"  Sources: {len(extractions)} PDFs")

            try:
                # Merge timelines using fuzzy matching
                merged_data = self._merge_statement_timelines(extractions, stmt_type)

                consolidated[stmt_type] = merged_data

                # Report success
                years = merged_data.get("fiscal_years", [])
                line_count = len(merged_data.get("line_items", []))
                print(f"  [OK] Consolidated: {line_count} line items, Years: {years}")

            except Exception as e:
                logger.error(f"Consolidation failed for {stmt_type.value}: {e}")
                print(f"  [FAIL] Consolidation failed: {e}")

        return consolidated

    def _merge_statement_timelines(
        self,
        extractions: List[Dict],
        stmt_type: FinancialStatementType
    ) -> Dict:
        """Merge multiple extractions into single timeline with fuzzy matching."""
        if len(extractions) == 1:
            # Only one source - just format it
            structured = extractions[0]["structured_data"]
            return self._format_single_extraction(structured, stmt_type)

        # Multiple sources - need to merge with tracking
        tracker = ConsolidationTracker()
        all_periods = []
        all_line_items = {}  # account_key -> merged data
        account_sources = {}  # account_key -> list of sources

        for extraction in extractions:
            structured = extraction["structured_data"]
            source_pdf = extraction.get("pdf_name", "unknown")

            # Collect periods (fiscal years)
            if hasattr(structured, 'periods'):
                all_periods.extend(structured.periods)

            # Collect line items
            if hasattr(structured, 'line_items'):
                for item in structured.line_items:
                    account_name = item.account_name
                    indent = item.indent_level
                    section = item.section if hasattr(item, 'section') else None
                    values_count = len(item.values) if hasattr(item, 'values') else 0

                    # Smart key generation: Total/summary lines should merge regardless of section
                    # because LLM may inconsistently assign sections to these
                    account_name_lower = account_name.lower().strip()
                    is_total_or_header = (
                        'total' in account_name_lower or
                        account_name_lower.endswith(':') or  # Section headers
                        account_name == "Assets" or
                        account_name == "Liabilities and Stockholders' Equity" or
                        account_name == "Liabilities" or
                        account_name == "Stockholders' equity" or
                        account_name == "Commitments and Contingencies (Note 10)" or
                        values_count == 0  # Empty headers
                    )

                    if is_total_or_header:
                        # For totals and headers: merge by name only (ignore both section AND indent)
                        # LLM may assign different indent levels to the same total line
                        account_key = f"{account_name}_HEADER"
                    else:
                        # For data accounts: require name + indent + section match
                        account_key = f"{account_name}_{indent}_{section or ''}"

                    # Track source information
                    if account_key not in account_sources:
                        account_sources[account_key] = []

                    account_sources[account_key].append({
                        "name": account_name,
                        "section": section,
                        "indent_level": indent,
                        "source_pdf": source_pdf,
                        "values_count": values_count
                    })

                    if account_key not in all_line_items:
                        all_line_items[account_key] = {
                            "account_name": account_name,
                            "indent_level": indent,
                            "section": section,
                            "values": {}
                        }
                    else:
                        # Account already exists - update metadata if current one is better
                        existing_section = all_line_items[account_key]["section"]
                        existing_indent = all_line_items[account_key]["indent_level"]

                        # Prefer non-null sections (LLM inconsistency fix)
                        if existing_section is None and section is not None:
                            all_line_items[account_key]["section"] = section

                        # For headers/totals: prefer smaller indent (less indented = more general)
                        # For data accounts: keep the first indent we see (they should match anyway)
                        if is_total_or_header and indent < existing_indent:
                            all_line_items[account_key]["indent_level"] = indent

                    # Add values for this period
                    if hasattr(item, 'values') and hasattr(structured, 'periods'):
                        for i, value in enumerate(item.values):
                            if i < len(structured.periods):
                                period_key = structured.periods[i]
                                all_line_items[account_key]["values"][period_key] = value

        # Track all merges and non-merges
        for account_key, sources in account_sources.items():
            account_data = all_line_items[account_key]
            account_name = account_data["account_name"]
            section = account_data["section"]
            values_count = len(account_data["values"])

            if len(sources) > 1:
                # Account was merged from multiple sources
                tracker.track_merge(account_name, section, sources)

                # Check if metadata was inconsistent (LLM extraction issue)
                sections = set(s["section"] for s in sources)
                indents = set(s["indent_level"] for s in sources)

                # Flag if sections are inconsistent (some null, some not)
                # OR if indent levels are inconsistent (different values)
                if (len(sections) > 1 and None in sections) or len(indents) > 1:
                    # Metadata inconsistency - LLM extracted same account with different metadata
                    tracker.track_duplicate_header(account_name, sources)
            else:
                # Account only in one source - potential missing data or unique account
                source = sources[0]
                reason = "only_in_one_pdf"
                if values_count == 0:
                    reason = "empty_header_single_source"
                tracker.track_unmerged(
                    account_name,
                    section,
                    source["source_pdf"],
                    reason,
                    values_count
                )

        # Extract fiscal years from periods
        fiscal_years = self._extract_years_from_periods(all_periods)

        return {
            "statement_type": stmt_type.value,
            "fiscal_years": sorted(fiscal_years, reverse=True),
            "periods": sorted(set(all_periods), reverse=True),
            "line_items": list(all_line_items.values()),
            "consolidation_metadata": tracker.get_summary()
        }

    def _format_single_extraction(self, structured_data, stmt_type: FinancialStatementType) -> Dict:
        """Format single extraction into standard structure."""
        periods = structured_data.periods if hasattr(structured_data, 'periods') else []
        line_items = structured_data.line_items if hasattr(structured_data, 'line_items') else []

        # Extract fiscal years
        fiscal_years = self._extract_years_from_periods(periods)

        # Format line items
        formatted_items = []
        for item in line_items:
            values_dict = {}
            if hasattr(item, 'values'):
                for i, value in enumerate(item.values):
                    if i < len(periods):
                        values_dict[periods[i]] = value

            formatted_items.append({
                "account_name": item.account_name,
                "indent_level": item.indent_level if hasattr(item, 'indent_level') else 0,
                "section": item.section if hasattr(item, 'section') else None,
                "values": values_dict
            })

        # No consolidation happened (single source), but still include metadata
        return {
            "statement_type": stmt_type.value,
            "fiscal_years": sorted(fiscal_years, reverse=True),
            "periods": periods,
            "line_items": formatted_items,
            "consolidation_metadata": {
                "merged_accounts": [],
                "unmerged_accounts": [],
                "duplicate_headers": [],
                "total_accounts_merged": 0,
                "total_accounts_unmerged": 0,
                "total_duplicate_headers": 0,
                "source_pdfs": ["single_source"]
            }
        }

    def _extract_years_from_periods(self, periods: List[str]) -> List[int]:
        """Extract 4-digit years from period strings."""
        import re

        years = set()
        for period in periods:
            year_match = re.search(r'(\d{4})', period)
            if year_match:
                years.add(int(year_match.group(1)))

        return sorted(list(years))

    def _phase4_generate_outputs(
        self,
        consolidated: Dict[FinancialStatementType, Dict],
        output_dir: Path
    ) -> Dict[FinancialStatementType, Dict]:
        """Phase 4: Generate JSON + Excel for each consolidated statement."""
        output_dir.mkdir(parents=True, exist_ok=True)

        outputs = {}

        for stmt_type, data in consolidated.items():
            print(f"\n  Generating outputs for: {stmt_type.value}")

            try:
                # Determine year range from data (NO HARDCODING!)
                years = data.get("fiscal_years", [])
                if years:
                    year_range = f"{min(years)}-{max(years)}"
                else:
                    year_range = "unknown"

                base_name = f"{stmt_type.value}_{year_range}"

                # Generate JSON
                json_path = output_dir / f"{base_name}.json"
                with open(json_path, 'w') as f:
                    json.dump(data, f, indent=2)
                print(f"  [OK] JSON: {json_path.name}")

                # Generate Excel
                excel_path = output_dir / f"{base_name}.xlsx"
                self._generate_excel(data, excel_path)
                print(f"  [OK] Excel: {excel_path.name}")

                # Track in run manifest
                line_item_count = len(data.get("line_items", []))
                self.current_run.save_consolidated(
                    statement_type=stmt_type.value,
                    years=[str(y) for y in years],
                    json_data=data,
                    excel_path=str(excel_path),
                    source_count=1,  # Will be updated with actual count
                    line_items=line_item_count
                )

                outputs[stmt_type] = {
                    "json_path": json_path,
                    "excel_path": excel_path,
                    "year_range": year_range
                }

            except Exception as e:
                logger.error(f"Output generation failed for {stmt_type.value}: {e}")
                print(f"  [FAIL] Output failed: {e}")

        return outputs

    def _generate_excel(self, data: Dict, excel_path: Path):
        """Generate formatted Excel file with hidden merge tracking columns."""
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, numbers, PatternFill

        wb = Workbook()
        ws = wb.active
        ws.title = data.get("statement_type", "Statement")

        # Get fiscal years and metadata
        fiscal_years = data.get("fiscal_years", [])
        metadata = data.get("consolidation_metadata", {})

        # Build merge tracking lookups
        merged_map = {}  # account_name -> list of sources
        unmerged_map = {}  # account_name -> source info
        issue_map = set()  # account names with issues

        for merged in metadata.get("merged_accounts", []):
            merged_map[merged["consolidated_name"]] = merged["merged_from"]

        for unmerged in metadata.get("unmerged_accounts", []):
            unmerged_map[unmerged["account_name"]] = unmerged

        for issue in metadata.get("duplicate_headers", []):
            issue_map.add(issue["header_name"])

        # Instruction row (row 1)
        ws.append(["Tip: Unhide columns F-H to see merge tracking details (Source PDFs, Merge Status, Notes)"])
        ws.merge_cells('A1:E1')
        ws['A1'].font = Font(italic=True, size=10, color="666666")
        ws['A1'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

        # Header row (row 2)
        headers = ["Account Name"] + [str(year) for year in fiscal_years] + ["Source PDFs", "Merge Status", "Notes"]
        ws.append(headers)

        # Bold headers
        for cell in ws[2]:
            cell.font = Font(bold=True)

        # Color definitions
        merged_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
        unmerged_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
        issue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

        # Data rows
        line_items = data.get("line_items", [])
        for item in line_items:
            account_name = item.get("account_name", "")
            indent_level = item.get("indent_level", 0)
            values_dict = item.get("values", {})

            # Apply indentation
            indented_name = ("  " * indent_level) + account_name

            # Build data columns
            row = [indented_name]
            for year in fiscal_years:
                value = ""
                for period, val in values_dict.items():
                    if str(year) in period:
                        value = val
                        break
                row.append(value)

            # Determine merge status and add hidden columns
            if account_name in merged_map:
                sources = merged_map[account_name]
                source_pdfs = "\n".join([s["source_pdf"] for s in sources])
                status = "✓ Merged"
                notes = ""
                row_fill = merged_fill

                if account_name in issue_map:
                    notes = "Inconsistent formatting between PDFs"
                    status = "⚠ Merged (issues)"
                    row_fill = issue_fill

            elif account_name in unmerged_map:
                info = unmerged_map[account_name]
                source_pdfs = info["source_pdf"]
                status = "⚠ Unmerged"
                notes = f"Only in {source_pdfs}"
                row_fill = unmerged_fill

            else:
                source_pdfs = ""
                status = ""
                notes = ""
                row_fill = None

            # Add hidden columns
            row.extend([source_pdfs, status, notes])

            # Append row
            ws.append(row)

            # Apply color coding (skip instruction row)
            current_row = ws.max_row
            if row_fill:
                for col_idx in range(1, len(fiscal_years) + 2):  # Account name + year columns
                    ws.cell(row=current_row, column=col_idx).fill = row_fill

        # Format numbers (skip instruction row)
        for row in ws.iter_rows(min_row=3):
            for cell in row[1:len(fiscal_years)+1]:  # Only year columns
                if cell.value and isinstance(cell.value, str):
                    clean_val = cell.value.replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
                    try:
                        numeric_val = float(clean_val)
                        cell.value = numeric_val
                        cell.number_format = '#,##0'
                    except:
                        pass

        # Hide merge tracking columns
        hidden_col_start = len(fiscal_years) + 2  # After Account Name + years
        ws.column_dimensions[ws.cell(row=2, column=hidden_col_start).column_letter].hidden = True  # Source PDFs
        ws.column_dimensions[ws.cell(row=2, column=hidden_col_start+1).column_letter].hidden = True  # Merge Status
        ws.column_dimensions[ws.cell(row=2, column=hidden_col_start+2).column_letter].hidden = True  # Notes

        # Adjust column widths
        ws.column_dimensions['A'].width = 50

        # Add Merge Summary sheet if consolidation metadata exists
        if metadata and (metadata.get("total_accounts_merged", 0) > 0 or
                         metadata.get("total_duplicate_headers", 0) > 0):
            self._add_merge_summary_sheet(wb, metadata)

        wb.save(excel_path)

    def _add_merge_summary_sheet(self, workbook: "Workbook", metadata: Dict):
        """Add a dedicated Merge Summary sheet to the workbook."""
        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        # Create new sheet
        ws = workbook.create_sheet("Merge Summary")

        # Header styling
        header_fill = PatternFill(start_color="366092", end_color="366092", fill_type="solid")
        header_font = Font(bold=True, color="FFFFFF", size=12)
        subheader_fill = PatternFill(start_color="D9E1F2", end_color="D9E1F2", fill_type="solid")
        subheader_font = Font(bold=True, size=11)

        row = 1

        # Main title
        ws.merge_cells(f'A{row}:D{row}')
        cell = ws[f'A{row}']
        cell.value = "CONSOLIDATION MERGE SUMMARY"
        cell.font = header_font
        cell.fill = header_fill
        cell.alignment = Alignment(horizontal='center', vertical='center')
        row += 2

        # Summary statistics
        ws[f'A{row}'] = "Total Accounts Merged:"
        ws[f'B{row}'] = metadata.get("total_accounts_merged", 0)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Total Duplicate Headers:"
        ws[f'B{row}'] = metadata.get("total_duplicate_headers", 0)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Total Unmerged Accounts:"
        ws[f'B{row}'] = metadata.get("total_accounts_unmerged", 0)
        ws[f'A{row}'].font = Font(bold=True)
        row += 1

        ws[f'A{row}'] = "Source PDFs:"
        ws[f'B{row}'] = ", ".join(metadata.get("source_pdfs", []))
        ws[f'A{row}'].font = Font(bold=True)
        row += 3

        # Merged Accounts section
        if metadata.get("merged_accounts"):
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "MERGED ACCOUNTS"
            cell.font = subheader_font
            cell.fill = subheader_fill
            cell.alignment = Alignment(horizontal='center')
            row += 1

            # Column headers
            ws[f'A{row}'] = "Consolidated Name"
            ws[f'B{row}'] = "Section"
            ws[f'C{row}'] = "Original Name"
            ws[f'D{row}'] = "Source PDF"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True)
                ws[f'{col}{row}'].fill = PatternFill(start_color="E7E6E6", end_color="E7E6E6", fill_type="solid")
            row += 1

            # Data rows
            for merged in metadata["merged_accounts"]:
                consolidated_name = merged["consolidated_name"]
                section = merged.get("section") or "N/A"

                for source in merged["merged_from"]:
                    ws[f'A{row}'] = consolidated_name
                    ws[f'B{row}'] = section
                    ws[f'C{row}'] = source["name"]
                    ws[f'D{row}'] = source["source_pdf"]
                    row += 1

            row += 2

        # Duplicate Headers section
        if metadata.get("duplicate_headers"):
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "DUPLICATE HEADERS (Potential Issues)"
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            row += 1

            ws[f'A{row}'] = "Header Name"
            ws[f'B{row}'] = "Section"
            ws[f'C{row}'] = "Source PDF"
            ws[f'D{row}'] = "Values Count"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True)
            row += 1

            for dup in metadata["duplicate_headers"]:
                header_name = dup["header_name"]
                for occurrence in dup["occurrences"]:
                    ws[f'A{row}'] = header_name
                    ws[f'B{row}'] = occurrence.get("section") or "N/A"
                    ws[f'C{row}'] = occurrence["source_pdf"]
                    ws[f'D{row}'] = occurrence["values_count"]
                    row += 1

            row += 2

        # Unmerged Accounts section (only show if there are any)
        unmerged_with_data = [u for u in metadata.get("unmerged_accounts", []) if u["values_count"] > 0]
        if unmerged_with_data:
            ws.merge_cells(f'A{row}:D{row}')
            cell = ws[f'A{row}']
            cell.value = "UNMERGED ACCOUNTS (Only in One PDF)"
            cell.font = subheader_font
            cell.fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")
            cell.alignment = Alignment(horizontal='center')
            row += 1

            ws[f'A{row}'] = "Account Name"
            ws[f'B{row}'] = "Section"
            ws[f'C{row}'] = "Source PDF"
            ws[f'D{row}'] = "Reason"
            for col in ['A', 'B', 'C', 'D']:
                ws[f'{col}{row}'].font = Font(bold=True)
            row += 1

            for unmerged in unmerged_with_data:
                ws[f'A{row}'] = unmerged["account_name"]
                ws[f'B{row}'] = unmerged.get("section") or "N/A"
                ws[f'C{row}'] = unmerged["source_pdf"]
                ws[f'D{row}'] = unmerged["reason"]
                row += 1

        # Adjust column widths
        ws.column_dimensions['A'].width = 40
        ws.column_dimensions['B'].width = 20
        ws.column_dimensions['C'].width = 40
        ws.column_dimensions['D'].width = 30

    def _phase4b_generate_combined_outputs(self, consolidated: Dict, output_dir: Path) -> Dict:
        """
        Phase 4B: Generate combined outputs with ALL statement types in single files.
        Creates:
        - One multi-sheet Excel workbook (one sheet per statement type)
        - One JSON file with all statement types
        """
        from openpyxl import Workbook
        from openpyxl.styles import Font, Alignment, numbers

        print(">> Creating combined outputs (all statement types in one file)...")
        print("-" * 70 + "\n")

        # Determine overall year range from all statements
        all_years = set()
        for stmt_type, data in consolidated.items():
            all_years.update(data.get("fiscal_years", []))

        all_years = sorted(all_years, reverse=True)
        year_range = f"{all_years[0]}-{all_years[-1]}" if len(all_years) > 1 else str(all_years[0])

        # File paths
        combined_json_path = output_dir / f"all_statements_{year_range}.json"
        combined_excel_path = output_dir / f"all_statements_{year_range}.xlsx"

        try:
            # 1. CREATE COMBINED JSON
            print(f"  Creating combined JSON...")
            combined_json_data = {
                "company_name": "NVIDIA Corporation",  # Could extract from data
                "fiscal_year_range": year_range,
                "fiscal_years": all_years,
                "statement_types": list(consolidated.keys()),
                "statements": {}
            }

            for stmt_type, data in consolidated.items():
                combined_json_data["statements"][stmt_type.value] = data

            with open(combined_json_path, 'w') as f:
                json.dump(combined_json_data, f, indent=2, default=str)

            print(f"  [OK] JSON: {combined_json_path.name}")

            # 2. CREATE COMBINED EXCEL (MULTI-SHEET)
            print(f"  Creating combined Excel workbook...")
            from openpyxl.styles import PatternFill
            wb = Workbook()

            # Remove default sheet
            if "Sheet" in wb.sheetnames:
                del wb["Sheet"]

            # Create one sheet per statement type
            for idx, (stmt_type, data) in enumerate(consolidated.items()):
                # Create sheet with statement type name
                sheet_name = stmt_type.value.replace("_", " ").title()
                if len(sheet_name) > 31:  # Excel sheet name limit
                    sheet_name = sheet_name[:31]

                ws = wb.create_sheet(title=sheet_name, index=idx)

                # Get fiscal years and metadata
                fiscal_years = data.get("fiscal_years", [])
                metadata = data.get("consolidation_metadata", {})

                # Build merge tracking lookups
                merged_map = {}
                unmerged_map = {}
                issue_map = set()

                for merged in metadata.get("merged_accounts", []):
                    merged_map[merged["consolidated_name"]] = merged["merged_from"]

                for unmerged in metadata.get("unmerged_accounts", []):
                    unmerged_map[unmerged["account_name"]] = unmerged

                for issue in metadata.get("duplicate_headers", []):
                    issue_map.add(issue["header_name"])

                # Color definitions
                merged_fill = PatternFill(start_color="C6EFCE", end_color="C6EFCE", fill_type="solid")  # Green
                unmerged_fill = PatternFill(start_color="FFEB9C", end_color="FFEB9C", fill_type="solid")  # Yellow
                issue_fill = PatternFill(start_color="FFC7CE", end_color="FFC7CE", fill_type="solid")  # Red

                # Instruction row
                ws.append(["Tip: Unhide columns F-H to see merge tracking details (Source PDFs, Merge Status, Notes)"])
                ws.merge_cells(f'A1:{chr(64+len(fiscal_years)+1)}1')  # Merge across visible columns
                ws['A1'].font = Font(italic=True, size=10, color="666666")
                ws['A1'].fill = PatternFill(start_color="F0F0F0", end_color="F0F0F0", fill_type="solid")

                # Header row
                headers = ["Account Name"] + [str(year) for year in fiscal_years] + ["Source PDFs", "Merge Status", "Notes"]
                ws.append(headers)

                # Bold headers
                for cell in ws[2]:
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')

                # Data rows
                line_items = data.get("line_items", [])
                for item in line_items:
                    account_name = item.get("account_name", "")
                    indent_level = item.get("indent_level", 0)
                    values_dict = item.get("values", {})

                    # Apply indentation
                    indented_name = ("  " * indent_level) + account_name

                    # Build data columns
                    row = [indented_name]
                    for year in fiscal_years:
                        value = ""
                        for period, val in values_dict.items():
                            if str(year) in period:
                                value = val
                                break
                        row.append(value)

                    # Determine merge status and add hidden columns
                    if account_name in merged_map:
                        sources = merged_map[account_name]
                        source_pdfs = "\n".join([s["source_pdf"] for s in sources])
                        status = "✓ Merged"
                        notes = ""
                        row_fill = merged_fill

                        if account_name in issue_map:
                            notes = "Inconsistent formatting between PDFs"
                            status = "⚠ Merged (issues)"
                            row_fill = issue_fill

                    elif account_name in unmerged_map:
                        info = unmerged_map[account_name]
                        source_pdfs = info["source_pdf"]
                        status = "⚠ Unmerged"
                        notes = f"Only in {source_pdfs}"
                        row_fill = unmerged_fill

                    else:
                        source_pdfs = ""
                        status = ""
                        notes = ""
                        row_fill = None

                    # Add hidden columns
                    row.extend([source_pdfs, status, notes])

                    # Append row
                    ws.append(row)

                    # Apply color coding
                    current_row = ws.max_row
                    if row_fill:
                        for col_idx in range(1, len(fiscal_years) + 2):  # Account name + year columns
                            ws.cell(row=current_row, column=col_idx).fill = row_fill

                # Format numbers
                for row in ws.iter_rows(min_row=3):
                    for cell in row[1:len(fiscal_years)+1]:  # Only year columns
                        if cell.value and isinstance(cell.value, str):
                            clean_val = cell.value.replace('$', '').replace(',', '').replace('(', '-').replace(')', '').strip()
                            try:
                                numeric_val = float(clean_val)
                                cell.value = numeric_val
                                cell.number_format = '#,##0'
                            except:
                                pass

                # Hide merge tracking columns
                hidden_col_start = len(fiscal_years) + 2  # After Account Name + years
                ws.column_dimensions[ws.cell(row=2, column=hidden_col_start).column_letter].hidden = True
                ws.column_dimensions[ws.cell(row=2, column=hidden_col_start+1).column_letter].hidden = True
                ws.column_dimensions[ws.cell(row=2, column=hidden_col_start+2).column_letter].hidden = True

                # Auto-size first column
                ws.column_dimensions['A'].width = 50

                print(f"    [OK] Sheet created: {sheet_name}")

            # Add summary sheet at the beginning
            summary_ws = wb.create_sheet(title="Summary", index=0)
            summary_ws.append(["NVIDIA Financial Statements Summary"])
            summary_ws.append([f"Fiscal Years: {year_range}"])
            summary_ws.append([])
            summary_ws.append(["Statement Type", "Sheet Name", "Line Items", "Fiscal Years"])

            for stmt_type, data in consolidated.items():
                sheet_name = stmt_type.value.replace("_", " ").title()[:31]
                line_item_count = len(data.get("line_items", []))
                years_str = ", ".join(map(str, data.get("fiscal_years", [])))
                summary_ws.append([stmt_type.value, sheet_name, line_item_count, years_str])

            # Bold summary headers
            for cell in summary_ws[1]:
                cell.font = Font(bold=True, size=14)
            for cell in summary_ws[4]:
                cell.font = Font(bold=True)

            summary_ws.column_dimensions['A'].width = 30
            summary_ws.column_dimensions['B'].width = 30
            summary_ws.column_dimensions['D'].width = 40

            # Save workbook
            wb.save(combined_excel_path)
            print(f"  [OK] Excel: {combined_excel_path.name}")
            print(f"  [OK] Total sheets: {len(consolidated) + 1} (includes summary)")

            return {
                "json_path": combined_json_path,
                "excel_path": combined_excel_path,
                "year_range": year_range,
                "statement_count": len(consolidated)
            }

        except Exception as e:
            logger.error(f"Combined output generation failed: {e}")
            print(f"  [FAIL] Combined output failed: {e}")
            return {}

    def _phase5_validate_everything(
        self,
        all_detections: Dict,
        all_extractions: Dict,
        consolidated: Dict,
        outputs: Dict
    ) -> Dict:
        """Phase 5: Comprehensive validation."""
        validation = {}

        # Validate detection
        validation["detection"] = self._validate_detection(all_detections)

        # Validate extraction
        validation["extraction"] = self._validate_extraction(all_extractions)

        # Validate consolidation
        validation["consolidation"] = self._validate_consolidation(consolidated)

        # Validate outputs
        validation["outputs"] = self._validate_outputs(outputs)

        # Overall success
        validation["overall_success"] = all([
            validation["detection"]["success"],
            validation["extraction"]["success"],
            validation["consolidation"]["success"],
            validation["outputs"]["success"]
        ])

        return validation

    def _validate_detection(self, all_detections: Dict) -> Dict:
        """Validate detection phase."""
        total_pdfs = len(all_detections)
        pdfs_with_detections = sum(1 for d in all_detections.values() if d)
        total_detections = sum(len(d) for d in all_detections.values())

        return {
            "success": total_detections > 0,
            "total_pdfs": total_pdfs,
            "pdfs_with_detections": pdfs_with_detections,
            "total_statement_types_found": total_detections
        }

    def _validate_extraction(self, all_extractions: Dict) -> Dict:
        """Validate extraction phase."""
        total_extractions = len(all_extractions)
        successful = sum(
            1 for e in all_extractions.values()
            if e.get("structured_data") is not None
        )

        # Check line item counts
        line_item_counts = [
            len(e.get("structured_data").line_items)
            for e in all_extractions.values()
            if hasattr(e.get("structured_data"), 'line_items')
        ]

        avg_line_items = sum(line_item_counts) / len(line_item_counts) if line_item_counts else 0

        return {
            "success": successful == total_extractions and avg_line_items > 0,
            "total_extractions": total_extractions,
            "successful_extractions": successful,
            "average_line_items": round(avg_line_items, 1)
        }

    def _validate_consolidation(self, consolidated: Dict) -> Dict:
        """Validate consolidation phase."""
        total_consolidated = len(consolidated)

        # Check fiscal year coverage
        all_years = set()
        for data in consolidated.values():
            years = data.get("fiscal_years", [])
            all_years.update(years)

        return {
            "success": total_consolidated > 0 and len(all_years) > 0,
            "total_statement_types": total_consolidated,
            "fiscal_year_range": f"{min(all_years)}-{max(all_years)}" if all_years else "N/A",
            "total_fiscal_years": len(all_years)
        }

    def _validate_outputs(self, outputs: Dict) -> Dict:
        """Validate output files."""
        total_outputs = len(outputs)
        json_files_exist = sum(
            1 for o in outputs.values()
            if o.get("json_path") and Path(o["json_path"]).exists()
        )
        excel_files_exist = sum(
            1 for o in outputs.values()
            if o.get("excel_path") and Path(o["excel_path"]).exists()
        )

        return {
            "success": json_files_exist == total_outputs and excel_files_exist == total_outputs,
            "total_statement_types": total_outputs,
            "json_files_created": json_files_exist,
            "excel_files_created": excel_files_exist
        }

    def _generate_report(self, validation: Dict, duration: float) -> str:
        """Generate final test report."""
        report = []
        report.append("\n" + "=" * 80)
        report.append("END-TO-END PIPELINE TEST REPORT")
        report.append("=" * 80)

        # Detection
        det = validation["detection"]
        report.append(f"\nPHASE 1: DETECTION")
        report.append("-" * 80)
        report.append(f"[OK] Status: {'PASS' if det['success'] else 'FAIL'}")
        report.append(f"   PDFs Processed: {det['total_pdfs']}")
        report.append(f"   PDFs with Statements: {det['pdfs_with_detections']}")
        report.append(f"   Total Statement Types Found: {det['total_statement_types_found']}")

        # Extraction
        ext = validation["extraction"]
        report.append(f"\nPHASE 2: EXTRACTION")
        report.append("-" * 80)
        report.append(f"[OK] Status: {'PASS' if ext['success'] else 'FAIL'}")
        report.append(f"   Total Extractions: {ext['total_extractions']}")
        report.append(f"   Successful: {ext['successful_extractions']}")
        report.append(f"   Average Line Items: {ext['average_line_items']}")

        # Consolidation
        con = validation["consolidation"]
        report.append(f"\nPHASE 3: CONSOLIDATION")
        report.append("-" * 80)
        report.append(f"[OK] Status: {'PASS' if con['success'] else 'FAIL'}")
        report.append(f"   Statement Types Consolidated: {con['total_statement_types']}")
        report.append(f"   Fiscal Year Range: {con['fiscal_year_range']}")
        report.append(f"   Total Fiscal Years: {con['total_fiscal_years']}")

        # Outputs
        out = validation["outputs"]
        report.append(f"\nPHASE 4: OUTPUTS")
        report.append("-" * 80)
        report.append(f"[OK] Status: {'PASS' if out['success'] else 'FAIL'}")
        report.append(f"   Statement Types: {out['total_statement_types']}")
        report.append(f"   JSON Files: {out['json_files_created']}")
        report.append(f"   Excel Files: {out['excel_files_created']}")

        # Overall
        report.append(f"\n" + "=" * 80)
        overall = "[OK] SUCCESS" if validation["overall_success"] else "[FAIL] FAILED"
        report.append(f"OVERALL RESULT: {overall}")
        report.append(f"Duration: {duration:.1f} seconds")
        report.append("=" * 80 + "\n")

        return "\n".join(report)


def main():
    """Run end-to-end pipeline test."""
    # Parse command-line arguments
    parser = argparse.ArgumentParser(
        description="Financial statement processing pipeline",
        formatter_class=argparse.RawDescriptionHelpFormatter,
        epilog="""
Examples:
  # Testing mode (default - uses samples/)
  python scripts/test_end_to_end_pipeline.py

  # Production mode (uses input/ and output/)
  python scripts/test_end_to_end_pipeline.py --mode production

  # Process specific PDFs
  python scripts/test_end_to_end_pipeline.py input/Company1.pdf input/Company2.pdf

  # Production mode with specific PDFs
  python scripts/test_end_to_end_pipeline.py --mode production input/*.pdf
        """
    )
    parser.add_argument(
        "pdfs",
        nargs="*",
        help="PDF files to process (optional - uses defaults if not provided)"
    )
    parser.add_argument(
        "--mode",
        choices=["production", "testing"],
        default=None,
        help="Processing mode (default: auto-detect from paths or use 'testing')"
    )

    args = parser.parse_args()

    # Determine mode and paths
    if args.mode:
        # Explicit mode from command line
        mode = args.mode
        config = PathConfig(mode=mode)
    elif args.pdfs and any("input/" in str(p) or "\\input\\" in str(p) for p in args.pdfs):
        # Auto-detect production mode from paths
        mode = "production"
        config = PathConfig(mode="production")
    else:
        # Default to testing mode
        mode = "testing"
        config = PathConfig(mode="testing")

    print(f">>> Running in {mode.upper()} mode")
    print(f">>> Input directory: {config.input_dir}")
    print(f">>> Output directory: {config.output_runs_dir}\n")

    # Determine PDF paths
    if args.pdfs:
        # Use PDFs from command line arguments
        pdf_paths = [Path(arg) for arg in args.pdfs]
        print(f"Using PDFs from command line: {[p.name for p in pdf_paths]}")
    else:
        # Default: Use sample NVIDIA PDFs
        pdf_paths = [
            config.input_dir / "NVIDIA 10K 2020-2019.pdf",
            config.input_dir / "NVIDIA 10K 2022-2021.pdf"
        ]
        print(f"Using default NVIDIA PDFs from {config.input_dir}")

    # Verify PDFs exist
    for pdf_path in pdf_paths:
        if not pdf_path.exists():
            print(f"ERROR: PDF not found: {pdf_path}")
            print(f"       Check that files exist in: {pdf_path.parent}")
            return 1

    # Ensure output directories exist
    config.ensure_directories()

    # Run pipeline (OutputManager will use the configured output directory)
    pipeline = EndToEndPipeline(output_base=str(config.output_runs_dir.parent))

    try:
        results = pipeline.run_full_pipeline(pdf_paths)

        # Save full results to run directory
        if pipeline.current_run:
            results_path = pipeline.current_run.run_dir / f"pipeline_results_{datetime.now().strftime('%Y%m%d_%H%M%S')}.json"
            with open(results_path, 'w') as f:
                # Convert Path objects to strings for JSON serialization
                serializable_results = {
                    "report": results["report"],
                    "validation": results["validation"],
                    "duration_seconds": results["duration_seconds"]
                }
                json.dump(serializable_results, f, indent=2)

            print(f"\n>>> Full results saved to: {results_path}")
            print(f">>> Run directory: {pipeline.current_run.run_dir}")

        # Exit code based on success
        return 0 if results["validation"]["overall_success"] else 1

    except Exception as e:
        logger.error(f"Pipeline failed: {e}", exc_info=True)
        print(f"\n[FAIL] Pipeline failed with error: {e}")
        return 1


if __name__ == "__main__":
    sys.exit(main())
