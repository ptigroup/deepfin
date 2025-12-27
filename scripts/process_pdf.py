#!/usr/bin/env python3
"""Process a single PDF through the financial document extraction pipeline.

This standalone CLI script processes a financial PDF and generates:
- JSON extraction with metadata
- Excel export with formatting

Usage:
    python scripts/process_pdf.py path/to/statement.pdf
    python scripts/process_pdf.py path/to/statement.pdf --output samples/output/
"""

import argparse
import asyncio
import json
import sys
from datetime import datetime
from pathlib import Path

# Add project root to path
sys.path.insert(0, str(Path(__file__).parent.parent))

from sqlalchemy.ext.asyncio import AsyncSession, async_sessionmaker, create_async_engine
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment

from app.core.config import get_settings
from app.core.database import AsyncSessionLocal, init_db
from app.core.logging import get_logger
from app.extraction.service import ExtractionService

logger = get_logger(__name__)


async def process_pdf(
    pdf_path: Path,
    output_dir: Path,
    company_name: str | None = None,
) -> dict:
    """Process a single PDF through the extraction pipeline.

    Args:
        pdf_path: Path to PDF file
        output_dir: Directory for output files
        company_name: Optional company name

    Returns:
        Processing result with metadata

    Raises:
        FileNotFoundError: If PDF doesn't exist
        Exception: If processing fails
    """
    if not pdf_path.exists():
        raise FileNotFoundError(f"PDF not found: {pdf_path}")

    print(f"\n{'='*70}")
    print(f"Processing: {pdf_path.name}")
    print(f"{'='*70}\n")

    start_time = datetime.now()

    # Initialize database
    await init_db()

    # Create session
    async with AsyncSessionLocal() as session:
        # Initialize extraction service
        extraction_service = ExtractionService(session)

        try:
            # Extract financial data
            print("→ Extracting financial data...")
            extraction_job = await extraction_service.extract_from_pdf(
                file_path=str(pdf_path),
                company_name=company_name or pdf_path.stem,
            )

            # Get extracted statement
            extracted_statement = extraction_job.statement

            if not extracted_statement:
                raise Exception("No statement extracted")

            print(f"✓ Extracted: {extracted_statement.statement_type.value}")
            print(f"  - Line items: {len(extracted_statement.line_items)}")
            print(f"  - Periods: {len(extracted_statement.periods)}")
            print(f"  - Company: {extracted_statement.company_name}")

            # Prepare metadata
            end_time = datetime.now()
            processing_time = (end_time - start_time).total_seconds()

            metadata = {
                "pdf_name": pdf_path.name,
                "pdf_size_kb": round(pdf_path.stat().st_size / 1024, 2),
                "document_type": extracted_statement.statement_type.value,
                "company_name": extracted_statement.company_name,
                "fiscal_year": extracted_statement.fiscal_year,
                "extracted_at": end_time.isoformat(),
                "processing_time_seconds": round(processing_time, 2),
                "periods_count": len(extracted_statement.periods),
                "line_items_count": len(extracted_statement.line_items),
                "extraction_mode": extraction_job.processing_mode or "direct",
                "status": extraction_job.status.value,
            }

            # Prepare output data
            output_data = {
                "metadata": metadata,
                "data": {
                    "company_name": extracted_statement.company_name,
                    "statement_type": extracted_statement.statement_type.value,
                    "fiscal_year": extracted_statement.fiscal_year,
                    "currency": extracted_statement.currency or "USD",
                    "periods": extracted_statement.periods,
                    "line_items": [
                        {
                            "account_name": item.account_name,
                            "values": item.values,
                            "indent_level": item.indent_level,
                            "parent_account": item.parent_account,
                            "category": item.category,
                        }
                        for item in extracted_statement.line_items
                    ],
                }
            }

            # Save JSON output
            json_dir = output_dir / "json"
            json_dir.mkdir(parents=True, exist_ok=True)
            json_path = json_dir / f"{pdf_path.stem}.json"

            with open(json_path, "w", encoding="utf-8") as f:
                json.dump(output_data, f, indent=2, ensure_ascii=False)

            print(f"\n✓ Saved JSON: {json_path}")

            # Save Excel output
            excel_dir = output_dir / "excel"
            excel_dir.mkdir(parents=True, exist_ok=True)
            excel_path = excel_dir / f"{pdf_path.stem}.xlsx"

            # Create Excel workbook
            wb = Workbook()
            ws = wb.active
            ws.title = extracted_statement.statement_type.value[:31]  # Excel max 31 chars

            # Header row
            header_font = Font(bold=True, size=12)
            center_align = Alignment(horizontal="center")
            right_align = Alignment(horizontal="right")

            # Write header
            ws.cell(1, 1, "Account").font = header_font
            for idx, period in enumerate(extracted_statement.periods, start=2):
                ws.cell(1, idx, period).font = header_font
                ws.cell(1, idx).alignment = center_align

            # Write line items
            for row_idx, item in enumerate(extracted_statement.line_items, start=2):
                # Apply indentation
                indent = "    " * item.indent_level
                ws.cell(row_idx, 1, f"{indent}{item.account_name}")

                # Write values
                for col_idx, value in enumerate(item.values, start=2):
                    cell = ws.cell(row_idx, col_idx, value)
                    cell.alignment = right_align

            # Auto-adjust column widths
            for col in ws.columns:
                max_length = 0
                column = col[0].column_letter
                for cell in col:
                    try:
                        if len(str(cell.value)) > max_length:
                            max_length = len(str(cell.value))
                    except:
                        pass
                adjusted_width = min(max_length + 2, 50)
                ws.column_dimensions[column].width = adjusted_width

            # Save workbook
            wb.save(excel_path)
            print(f"✓ Saved Excel: {excel_path}")

            # Save metadata separately
            metadata_dir = output_dir / "metadata"
            metadata_dir.mkdir(parents=True, exist_ok=True)
            metadata_path = metadata_dir / f"{pdf_path.stem}_metadata.json"

            with open(metadata_path, "w", encoding="utf-8") as f:
                json.dump(metadata, f, indent=2)

            print(f"✓ Saved metadata: {metadata_path}")

            print(f"\n{'='*70}")
            print(f"✓ SUCCESS - Completed in {processing_time:.2f} seconds")
            print(f"{'='*70}\n")

            return {
                "status": "success",
                "pdf": pdf_path.name,
                "metadata": metadata,
                "outputs": {
                    "json": str(json_path),
                    "excel": str(excel_path),
                    "metadata": str(metadata_path),
                }
            }

        except Exception as e:
            print(f"\n✗ ERROR: {str(e)}")
            logger.error(
                "PDF processing failed",
                extra={"pdf": pdf_path.name, "error": str(e)},
                exc_info=True
            )

            return {
                "status": "error",
                "pdf": pdf_path.name,
                "error": str(e),
            }


async def main():
    """Main entry point."""
    parser = argparse.ArgumentParser(
        description="Process financial PDF through extraction pipeline"
    )
    parser.add_argument(
        "pdf_path",
        type=Path,
        help="Path to PDF file"
    )
    parser.add_argument(
        "--output",
        "-o",
        type=Path,
        default=Path("samples/output"),
        help="Output directory (default: samples/output)"
    )
    parser.add_argument(
        "--company",
        "-c",
        type=str,
        help="Company name (default: inferred from filename)"
    )

    args = parser.parse_args()

    try:
        result = await process_pdf(
            pdf_path=args.pdf_path,
            output_dir=args.output,
            company_name=args.company,
        )

        # Exit with appropriate code
        sys.exit(0 if result["status"] == "success" else 1)

    except KeyboardInterrupt:
        print("\n\nInterrupted by user")
        sys.exit(130)
    except Exception as e:
        print(f"\n\nFatal error: {str(e)}")
        logger.error("Fatal error in main", exc_info=True)
        sys.exit(1)


if __name__ == "__main__":
    asyncio.run(main())
