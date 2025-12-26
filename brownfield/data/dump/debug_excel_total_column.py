#!/usr/bin/env python3

import json
import pandas as pd
from openpyxl import load_workbook

def debug_excel_total_column():
    """Debug why Total Shareholders' Equity column has no values in Excel."""
    
    print("🔍 DEBUGGING EXCEL TOTAL SHAREHOLDERS' EQUITY COLUMN")
    print("=" * 60)
    
    # First, check the JSON data to confirm values are there
    print("📊 STEP 1: Check JSON data")
    with open('/mnt/c/Claude/LLMWhisperer/output/structured/json/shareholder equity_schema_based_extraction.json', 'r') as f:
        data = json.load(f)
    
    structured_data = json.loads(data['structured_data'])
    
    # Look for Total Shareholders' Equity values in JSON
    total_equity_count = 0
    missing_count = 0
    
    for row in structured_data['equity_rows']:
        total_value = row['column_values'].get("Total Shareholders' Equity", "MISSING")
        if total_value and total_value != "MISSING" and total_value.strip():
            total_equity_count += 1
            if row['row_type'] == 'balance':
                print(f"  ✅ {row['transaction_description'][:30]:30} | Total: {total_value}")
        else:
            missing_count += 1
            if row['row_type'] == 'balance':
                print(f"  ❌ {row['transaction_description'][:30]:30} | Total: MISSING")
    
    print(f"\n📈 JSON Summary: {total_equity_count} rows have total values, {missing_count} rows are missing")
    
    # Step 2: Check what the Excel file actually contains
    print(f"\n📊 STEP 2: Check Excel file content")
    
    try:
        # Read Excel file to see the actual data
        df = pd.read_excel('/mnt/c/Claude/LLMWhisperer/output/structured/excel/shareholder equity_schema_based_extraction.xlsx')
        
        print(f"Excel file loaded successfully")
        print(f"Columns: {list(df.columns)}")
        
        # Check for Total Shareholders' Equity column
        total_col = None
        for col in df.columns:
            if "Total" in str(col) and "Equity" in str(col):
                total_col = col
                break
        
        if total_col:
            print(f"Found total column: '{total_col}'")
            non_empty_values = df[total_col].dropna()
            non_empty_count = len(non_empty_values)
            print(f"Non-empty values in total column: {non_empty_count}")
            
            # Show some sample values
            if non_empty_count > 0:
                print(f"Sample values:")
                for i, value in enumerate(non_empty_values.head()):
                    print(f"  {i+1}: '{value}'")
            else:
                print("❌ ALL VALUES IN TOTAL COLUMN ARE EMPTY!")
                
                # Let's check all columns to see where data might be going
                print(f"\nChecking all columns for data:")
                for col in df.columns:
                    non_empty = len(df[col].dropna())
                    print(f"  {col}: {non_empty} non-empty values")
        else:
            print("❌ No Total Shareholders' Equity column found in Excel!")
            print(f"Available columns: {list(df.columns)}")
            
    except Exception as e:
        print(f"❌ Error reading Excel file: {e}")
        
        # Try using openpyxl directly
        try:
            print(f"\nTrying with openpyxl...")
            wb = load_workbook('/mnt/c/Claude/LLMWhisperer/output/structured/excel/shareholder equity_schema_based_extraction.xlsx')
            ws = wb.active
            
            print(f"Worksheet loaded: {ws.title}")
            print(f"Max row: {ws.max_row}, Max col: {ws.max_column}")
            
            # Check header row
            headers = []
            for col in range(1, ws.max_column + 1):
                header = ws.cell(1, col).value
                headers.append(header)
                print(f"Col {col}: '{header}'")
                
            # Find Total Shareholders' Equity column
            total_col_idx = None
            for i, header in enumerate(headers):
                if header and "Total" in str(header) and "Equity" in str(header):
                    total_col_idx = i + 1
                    break
            
            if total_col_idx:
                print(f"\nFound Total Shareholders' Equity in column {total_col_idx}")
                print(f"Sample values from column {total_col_idx}:")
                for row in range(2, min(8, ws.max_row + 1)):  # Check first few data rows
                    value = ws.cell(row, total_col_idx).value
                    desc = ws.cell(row, 1).value if ws.max_column > 0 else "N/A"
                    print(f"  Row {row}: '{desc}' | Total: '{value}'")
            else:
                print(f"❌ Total Shareholders' Equity column not found in Excel headers")
                
        except Exception as e2:
            print(f"❌ Error with openpyxl: {e2}")

if __name__ == "__main__":
    debug_excel_total_column()