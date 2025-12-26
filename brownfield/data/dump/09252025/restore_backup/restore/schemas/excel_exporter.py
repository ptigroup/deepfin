"""
Schema-Based Excel Exporter for Financial Statements.

This module provides a generic Excel exporter that reads schema layout configuration
and creates properly formatted Excel files that match the original table structure.
"""

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, Border, Side, PatternFill
from openpyxl.utils import get_column_letter
from datetime import datetime
from typing import Dict, Any

from .base_schema import BaseFinancialSchema, ExcelLayoutConfig, ExcelColumnMapping

class SchemaBasedExcelExporter:
    """Generic Excel exporter that uses schema layout configuration."""
    
    def __init__(self):
        """Initialize the exporter."""
        self.workbook = None
        self.worksheet = None
        
    def export_to_excel(self, schema_instance: BaseFinancialSchema, output_path: str) -> None:
        """
        Export schema instance to Excel using its layout configuration.
        
        Args:
            schema_instance: Instance of a financial schema
            output_path: Path where Excel file should be saved
        """
        # Get layout configuration from schema
        layout_config = schema_instance.get_excel_layout_config()
        
        # Create workbook and worksheet
        self.workbook = Workbook()
        self.worksheet = self.workbook.active
        self.worksheet.title = "Financial Statement"
        
        # Build the Excel file
        current_row = 1
        
        # Add header rows (company name, document title, units)
        current_row = self._add_header_rows(layout_config, current_row)
        
        # Add table headers (main + sub headers if applicable)
        current_row = self._add_table_headers(layout_config, current_row)
        
        # Add data rows
        current_row = self._add_data_rows(schema_instance, layout_config, current_row)
        
        # Add units note at bottom if specified
        if layout_config.units_note_position == "bottom" and hasattr(schema_instance, 'units_note') and schema_instance.units_note:
            current_row = self._add_units_note_at_bottom(schema_instance, layout_config, current_row)
        
        # Apply formatting
        self._apply_formatting(layout_config)
        
        # Save the file
        self.workbook.save(output_path)
        print(f"✅ Excel saved to: {output_path}")
    
    def _add_header_rows(self, layout_config: ExcelLayoutConfig, start_row: int) -> int:
        """Add header rows (company name, title, units) to Excel."""
        current_row = start_row
        
        for header_text in layout_config.header_rows:
            if header_text.strip():  # Only add non-empty headers
                # Merge across all columns and center
                max_col = max(len(layout_config.column_mappings) + 1, 7)  # At least 7 columns
                self.worksheet.merge_cells(f'A{current_row}:{get_column_letter(max_col)}{current_row}')
                
                cell = self.worksheet[f'A{current_row}']
                cell.value = header_text
                cell.font = Font(bold=True, size=12)
                cell.alignment = Alignment(horizontal='center')
                current_row += 1
        
        # Add blank row after headers
        current_row += 1
        return current_row
    
    def _add_table_headers(self, layout_config: ExcelLayoutConfig, start_row: int) -> int:
        """Add table column headers to Excel."""
        if not layout_config.column_mappings:
            return start_row
        
        current_row = start_row
        
        # Add transaction description header in column A
        self.worksheet['A' + str(current_row)].value = ""
        
        if layout_config.has_multi_level_headers:
            # Add main headers row
            main_headers_added = set()
            for mapping in layout_config.column_mappings:
                col_letter = get_column_letter(mapping.excel_column_index)
                
                # Only add main header if we haven't seen it before
                if mapping.main_header not in main_headers_added:
                    cell = self.worksheet[col_letter + str(current_row)]
                    cell.value = mapping.main_header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
                    
                    # Handle merging for multi-column headers like "Common Stock Outstanding"
                    if mapping.merge_with_next:
                        # Find the next column for this main header
                        next_col_index = mapping.excel_column_index + 1
                        if next_col_index <= len(layout_config.column_mappings) + 1:
                            next_col_letter = get_column_letter(next_col_index)
                            self.worksheet.merge_cells(f'{col_letter}{current_row}:{next_col_letter}{current_row}')
                    
                    main_headers_added.add(mapping.main_header)
            
            current_row += 1
            
            # Add sub-headers row
            for mapping in layout_config.column_mappings:
                if mapping.sub_header:  # Only add if there's a sub-header
                    col_letter = get_column_letter(mapping.excel_column_index)
                    cell = self.worksheet[col_letter + str(current_row)]
                    cell.value = mapping.sub_header
                    cell.font = Font(bold=True)
                    cell.alignment = Alignment(horizontal='center')
        else:
            # Single level headers
            for mapping in layout_config.column_mappings:
                col_letter = get_column_letter(mapping.excel_column_index)
                cell = self.worksheet[col_letter + str(current_row)]
                cell.value = mapping.main_header
                cell.font = Font(bold=True)
                cell.alignment = Alignment(horizontal='center')
        
        current_row += 1
        return current_row
    
    def _add_data_rows(self, schema_instance: BaseFinancialSchema, layout_config: ExcelLayoutConfig, start_row: int) -> int:
        """Add data rows based on the schema type."""
        current_row = start_row
        
        # Handle different schema types
        if hasattr(schema_instance, 'equity_rows'):
            # Shareholders equity schema
            for row in schema_instance.equity_rows:
                # Add transaction description in column A
                self.worksheet[f'A{current_row}'].value = row.transaction_description
                
                # Add values for each column
                for mapping in layout_config.column_mappings:
                    col_letter = get_column_letter(mapping.excel_column_index)
                    
                    # Create the key to look up the value
                    if mapping.sub_header:
                        key = f"{mapping.main_header}:{mapping.sub_header}"
                    else:
                        # For single headers, try both patterns:
                        # 1. Without colon (e.g., "Total Shareholders' Equity")
                        # 2. With colon (e.g., "Additional Paid-in Capital:")
                        key_without_colon = mapping.main_header
                        key_with_colon = f"{mapping.main_header}:"
                        
                        if key_without_colon in row.column_values:
                            key = key_without_colon
                        elif key_with_colon in row.column_values:
                            key = key_with_colon
                        else:
                            key = key_without_colon  # Default to without colon
                    
                    # Get value from column_values
                    value = row.column_values.get(key, "")
                    if value and value != "-":
                        self.worksheet[col_letter + str(current_row)].value = value
                
                current_row += 1
        
        elif hasattr(schema_instance, 'accounts'):
            # Balance sheet schema with accounts
            for item in schema_instance.accounts:
                # Add account name in column A with proper indentation
                account_name = item.account_name
                
                # Apply visual indentation based on indent_level
                if hasattr(item, 'indent_level') and item.indent_level > 0:
                    # Add spaces for indentation (4 spaces per level)
                    indent_spaces = "    " * item.indent_level
                    account_name = indent_spaces + account_name
                
                self.worksheet[f'A{current_row}'].value = account_name
                
                # Add values for each period
                if hasattr(item, 'values') and item.values:
                    col_index = 2  # Start from column B
                    for period, value in item.values.items():
                        if col_index <= len(layout_config.column_mappings) + 1:
                            col_letter = get_column_letter(col_index)
                            self.worksheet[col_letter + str(current_row)].value = value
                            col_index += 1
                
                current_row += 1
        
        elif hasattr(schema_instance, 'line_items'):
            # Other schema types with line_items
            for item in schema_instance.line_items:
                # Add account name in column A with proper indentation
                account_name = item.account_name
                
                # Apply visual indentation based on indent_level
                if hasattr(item, 'indent_level') and item.indent_level > 0:
                    # Add spaces for indentation (4 spaces per level)
                    indent_spaces = "    " * item.indent_level
                    account_name = indent_spaces + account_name
                
                self.worksheet[f'A{current_row}'].value = account_name
                
                # Add values for each period
                if hasattr(item, 'values') and item.values:
                    col_index = 2  # Start from column B
                    for period, value in item.values.items():
                        if col_index <= len(layout_config.column_mappings) + 1:
                            col_letter = get_column_letter(col_index)
                            self.worksheet[col_letter + str(current_row)].value = value
                            col_index += 1
                
                current_row += 1
        
        return current_row
    
    def _add_units_note_at_bottom(self, schema_instance: BaseFinancialSchema, layout_config: ExcelLayoutConfig, start_row: int) -> int:
        """Add units note at the bottom of the table with 📊 emoji."""
        current_row = start_row + 1  # Add blank row before units note
        
        # Create units note with 📊 emoji
        units_text = f"📊 UNITS: {schema_instance.units_note}"
        
        # Merge across all columns
        max_col = max(len(layout_config.column_mappings) + 1, 7)
        self.worksheet.merge_cells(f'A{current_row}:{get_column_letter(max_col)}{current_row}')
        
        cell = self.worksheet[f'A{current_row}']
        cell.value = units_text
        cell.font = Font(bold=True, size=10)
        cell.alignment = Alignment(horizontal='left')
        
        return current_row + 1
    
    def _apply_formatting(self, layout_config: ExcelLayoutConfig) -> None:
        """Apply formatting to the Excel worksheet."""
        # Define styles
        thin_border = Border(
            left=Side(style='thin'),
            right=Side(style='thin'), 
            top=Side(style='thin'),
            bottom=Side(style='thin')
        )
        
        # Apply borders to data area
        if layout_config.column_mappings:
            max_col = len(layout_config.column_mappings) + 1
            max_row = self.worksheet.max_row
            
            # Apply borders to the entire table area
            start_row = layout_config.table_start_row
            for row in range(start_row, max_row + 1):
                for col in range(1, max_col + 1):
                    cell = self.worksheet.cell(row=row, column=col)
                    cell.border = thin_border
                    
                    # Apply alignment based on data type
                    if col == 1:  # Transaction description column
                        cell.alignment = Alignment(horizontal='left')
                    else:
                        cell.alignment = Alignment(horizontal='right')
        
        # Auto-adjust column widths
        for column in self.worksheet.columns:
            max_length = 0
            column_letter = get_column_letter(column[0].column)
            
            for cell in column:
                if cell.value:
                    max_length = max(max_length, len(str(cell.value)))
            
            adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
            self.worksheet.column_dimensions[column_letter].width = adjusted_width