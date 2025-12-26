#!/usr/bin/env python3

import json
from schemas.excel_exporter import SchemaBasedExcelExporter
from schemas import schema_registry
from schemas.base_schema import FinancialStatementType
from openpyxl import load_workbook

def test_excel_debug():
    """Test Excel export with detailed debugging."""
    
    print("🔍 DEBUGGING EXCEL EXPORT")
    print("=" * 50)
    
    # Load the JSON data
    with open('/mnt/c/Claude/LLMWhisperer/output/structured/json/shareholder equity_schema_based_extraction.json', 'r') as f:
        data = json.load(f)
    
    # Parse the structured data
    financial_data = json.loads(data['structured_data'])
    
    # Get schema class
    schema_class = schema_registry.get_schema_class(FinancialStatementType.SHAREHOLDERS_EQUITY)
    schema_instance = schema_class.model_validate(financial_data)
    
    # Use the schema-based Excel exporter
    output_path = 'output/structured/excel/debug_export.xlsx'
    exporter = SchemaBasedExcelExporter()
    
    print(f"🚀 Exporting to: {output_path}")
    exporter.export_to_excel(schema_instance, output_path)
    
    # Now let's inspect the Excel file directly using openpyxl
    print(f"\n🔍 INSPECTING EXCEL FILE:")
    workbook = load_workbook(output_path)
    worksheet = workbook.active
    
    print(f"📊 WORKSHEET DIMENSIONS: {worksheet.max_row} rows x {worksheet.max_column} cols")
    
    # Show first 5 rows to see headers and data
    print(f"\n📋 FIRST 5 ROWS:")
    for row_num in range(1, min(6, worksheet.max_row + 1)):
        row_data = []
        for col_num in range(1, worksheet.max_column + 1):
            cell_value = worksheet.cell(row=row_num, column=col_num).value
            row_data.append(str(cell_value) if cell_value is not None else "")
        print(f"  Row {row_num}: {row_data}")
    
    # Now test pandas reading
    import pandas as pd
    print(f"\n🐼 PANDAS READING:")
    df = pd.read_excel(output_path)
    print(f"Columns: {list(df.columns)}")
    print(f"Shape: {df.shape}")
    
    # Look for Total column by value pattern instead of name
    print(f"\n🔍 SEARCHING FOR TOTAL VALUES:")
    for col_idx, col_name in enumerate(df.columns):
        col_data = df[col_name].dropna()
        # Look for dollar signs and large numbers (typical of totals)
        total_like_values = [val for val in col_data if isinstance(val, str) and ('$' in val or ',' in val) and any(char.isdigit() for char in val)]
        if len(total_like_values) > 5:  # If many values look like totals
            print(f"  Column '{col_name}' has {len(total_like_values)} total-like values:")
            for val in total_like_values[:5]:  # Show first 5
                print(f"    {val}")
    
    workbook.close()

if __name__ == "__main__":
    test_excel_debug()