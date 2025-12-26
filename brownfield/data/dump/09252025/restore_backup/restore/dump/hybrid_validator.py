#!/usr/bin/env python3
"""
Hybrid Financial Document Validator
Combines traditional numerical validation with GPT-4o-mini semantic validation

Pipeline: Traditional Numerical Validation (70%) + GPT-4o-mini Semantic Validation (30%) = Final Score
"""

import os
import re
import json
import pdfplumber
import pandas as pd
from pathlib import Path
from dotenv import load_dotenv
from openai import OpenAI
from rich.console import Console
from rich.table import Table
from rich.panel import Panel
from rich.progress import track
from tabulate import tabulate
from openpyxl import load_workbook

class HybridValidator:
    def __init__(self):
        """Initialize the hybrid validator with both traditional and LLM capabilities."""
        # Load environment variables
        load_dotenv()
        
        # Initialize OpenAI client
        self.openai_client = OpenAI(
            api_key=os.getenv('OPENAI_API_KEY'),
            base_url=os.getenv('OPENAI_BASE_URL', 'https://api.openai.com/v1')
        )
        
        # Initialize Rich console for beautiful output
        self.console = Console()
        
        # Validation results storage
        self.results = {
            'traditional': {},
            'semantic': {},
            'hybrid': {}
        }
        
    def validate_extraction(self, pdf_path, excel_path, llm_text=None):
        """
        Run complete hybrid validation on PDF → Excel extraction.
        
        Args:
            pdf_path (str): Path to original PDF file
            excel_path (str): Path to generated Excel file
            llm_text (str, optional): Pre-extracted text from LLMWhisperer
            
        Returns:
            dict: Complete validation results with scores and details
        """
        self.console.print("\n🔍 [bold blue]Starting Hybrid Validation Pipeline[/bold blue]")
        self.console.print("=" * 60)
        
        # Phase 1: Traditional Numerical Validation
        self.console.print("\n📊 [yellow]Phase 1: Traditional Numerical Validation[/yellow]")
        traditional_score = self._traditional_validation(pdf_path, excel_path, llm_text)
        
        # Phase 2: GPT-4o-mini Semantic Validation  
        self.console.print("\n🤖 [cyan]Phase 2: GPT-4o-mini Semantic Validation[/cyan]")
        semantic_score = self._semantic_validation(pdf_path, excel_path, llm_text)
        
        # Phase 3: Hybrid Scoring
        self.console.print("\n⚖️  [green]Phase 3: Hybrid Scoring[/green]")
        final_score = self._calculate_hybrid_score(traditional_score, semantic_score)
        
        # Phase 4: Generate Report
        self.console.print("\n📋 [magenta]Phase 4: Validation Report[/magenta]")
        self._generate_report()
        
        return self.results
    
    def _traditional_validation(self, pdf_path, excel_path, llm_text=None):
        """Traditional numerical validation using exact comparison."""
        self.console.print("🔢 Extracting numerical data from source...")
        
        # Extract numbers from PDF or LLM text
        if llm_text:
            self.console.print("📄 Using LLMWhisperer extracted text for analysis")
            pdf_numbers = self._extract_numbers_from_text(llm_text)
        else:
            pdf_numbers = self._extract_pdf_numbers(pdf_path)
        
        self.console.print("📈 Extracting numerical data from Excel...")
        
        # Extract numbers from Excel
        excel_numbers = self._extract_excel_numbers(excel_path)
        
        self.console.print("🔍 Comparing numerical accuracy...")
        
        # Compare numerical data
        comparison = self._compare_numbers(pdf_numbers, excel_numbers)
        
        # Calculate traditional score
        if comparison['total_pdf_numbers'] > 0:
            score = (comparison['matched_numbers'] / comparison['total_pdf_numbers']) * 100
        else:
            score = 0
            
        self.results['traditional'] = {
            'score': round(score, 2),
            'pdf_numbers': pdf_numbers,
            'excel_numbers': excel_numbers,
            'comparison': comparison
        }
        
        self.console.print(f"✅ Traditional validation score: [bold green]{score:.1f}%[/bold green]")
        return score
    
    def _extract_pdf_numbers(self, pdf_path):
        """Extract all numerical values from PDF using pdfplumber."""
        numbers = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        # Extract numbers with various formats
                        number_patterns = [
                            r'\$\s*([0-9,]+\.?\d*)',  # $10,896 or $10,896.50
                            r'\b([0-9,]+\.?\d*)\b',   # 10,896 or 10896.50
                            r'\(\s*([0-9,]+\.?\d*)\s*\)',  # (10,896) negative numbers
                        ]
                        
                        for pattern in number_patterns:
                            matches = re.findall(pattern, text)
                            for match in matches:
                                # Clean and convert to float
                                clean_num = re.sub(r'[,\s]', '', match)
                                try:
                                    num = float(clean_num)
                                    if num > 0:  # Only positive numbers for financial data
                                        numbers.append(num)
                                except ValueError:
                                    continue
                                    
        except Exception as e:
            self.console.print(f"⚠️  Error extracting PDF numbers: {e}")
            
        # Remove duplicates and sort
        unique_numbers = list(set(numbers))
        unique_numbers.sort(reverse=True)  # Largest first
        
        return unique_numbers[:50]  # Top 50 numbers to avoid noise
    
    def _extract_numbers_from_text(self, text):
        """Extract numerical values from LLMWhisperer text."""
        numbers = []
        
        if not text:
            return numbers
            
        # Extract numbers with various formats
        number_patterns = [
            r'\$\s*([0-9,]+\.?\d*)',  # $10,896 or $10,896.50
            r'\b([0-9,]+\.?\d*)\b',   # 10,896 or 10896.50
            r'\(\s*([0-9,]+\.?\d*)\s*\)',  # (10,896) negative numbers
        ]
        
        for pattern in number_patterns:
            matches = re.findall(pattern, text)
            for match in matches:
                # Clean and convert to float
                clean_num = re.sub(r'[,\s]', '', match)
                try:
                    num = float(clean_num)
                    if num > 0:  # Only positive numbers for financial data
                        numbers.append(num)
                except ValueError:
                    continue
        
        # Remove duplicates and sort
        unique_numbers = list(set(numbers))
        unique_numbers.sort(reverse=True)  # Largest first
        
        return unique_numbers[:50]  # Top 50 numbers
    
    def _extract_excel_numbers(self, excel_path):
        """Extract all numerical values from Excel file."""
        numbers = []
        
        try:
            wb = load_workbook(excel_path)
            ws = wb['Balance Sheet']  # Main sheet
            
            for row in ws.iter_rows():
                for cell in row:
                    if cell.value is not None:
                        # Try to convert to number
                        if isinstance(cell.value, (int, float)):
                            if cell.value > 0:
                                numbers.append(float(cell.value))
                        elif isinstance(cell.value, str):
                            # Extract numbers from strings
                            clean_str = re.sub(r'[,$\s]', '', str(cell.value))
                            try:
                                num = float(clean_str)
                                if num > 0:
                                    numbers.append(num)
                            except ValueError:
                                continue
                                
        except Exception as e:
            self.console.print(f"⚠️  Error extracting Excel numbers: {e}")
            
        # Remove duplicates and sort
        unique_numbers = list(set(numbers))
        unique_numbers.sort(reverse=True)
        
        return unique_numbers[:50]  # Top 50 numbers
    
    def _compare_numbers(self, pdf_numbers, excel_numbers):
        """Compare numerical data between PDF and Excel."""
        matched = 0
        missing_from_excel = []
        extra_in_excel = []
        
        # Convert to sets for comparison
        pdf_set = set(pdf_numbers)
        excel_set = set(excel_numbers)
        
        # Find matches
        matches = pdf_set.intersection(excel_set)
        matched = len(matches)
        
        # Find missing and extra numbers
        missing_from_excel = list(pdf_set - excel_set)
        extra_in_excel = list(excel_set - pdf_set)
        
        return {
            'matched_numbers': matched,
            'total_pdf_numbers': len(pdf_numbers),
            'total_excel_numbers': len(excel_numbers),
            'missing_from_excel': missing_from_excel[:10],  # Top 10 missing
            'extra_in_excel': extra_in_excel[:10],  # Top 10 extra
            'accuracy_percentage': (matched / len(pdf_numbers) * 100) if pdf_numbers else 0
        }
    
    def _semantic_validation(self, pdf_path, excel_path, llm_text=None):
        """Semantic validation using GPT-4o-mini."""
        self.console.print("📄 Preparing content for semantic analysis...")
        
        # Extract text content
        if llm_text:
            pdf_text = llm_text[:3000]  # Use LLMWhisperer text
        else:
            pdf_text = self._extract_pdf_text(pdf_path)
            
        excel_content = self._extract_excel_content(excel_path)
        
        self.console.print("🤖 Sending to GPT-4o-mini for semantic analysis...")
        
        # Create validation prompt
        prompt = self._create_validation_prompt(pdf_text, excel_content)
        
        try:
            # Call GPT-4o-mini
            response = self.openai_client.chat.completions.create(
                model=os.getenv('OPENAI_CHOICE', 'gpt-4o-mini'),
                messages=[
                    {"role": "system", "content": "You are a financial document validation expert. Analyze the extraction accuracy between a PDF and Excel file."},
                    {"role": "user", "content": prompt}
                ],
                temperature=0.1,  # Low temperature for consistent results
                max_tokens=2000
            )
            
            # Parse GPT response
            gpt_analysis = response.choices[0].message.content
            semantic_score = self._parse_gpt_score(gpt_analysis)
            
            self.results['semantic'] = {
                'score': semantic_score,
                'gpt_analysis': gpt_analysis,
                'model_used': os.getenv('OPENAI_CHOICE', 'gpt-4o-mini')
            }
            
            self.console.print(f"✅ Semantic validation score: [bold cyan]{semantic_score:.1f}%[/bold cyan]")
            return semantic_score
            
        except Exception as e:
            self.console.print(f"⚠️  Error in GPT validation: {e}")
            # Fallback score
            semantic_score = 75.0
            self.results['semantic'] = {
                'score': semantic_score,
                'error': str(e),
                'gpt_analysis': "Error occurred during GPT analysis"
            }
            return semantic_score
    
    def _extract_pdf_text(self, pdf_path):
        """Extract clean text from PDF for semantic analysis."""
        text_content = []
        
        try:
            with pdfplumber.open(pdf_path) as pdf:
                for page in pdf.pages:
                    text = page.extract_text()
                    if text:
                        text_content.append(text)
                        
        except Exception as e:
            self.console.print(f"⚠️  Error extracting PDF text: {e}")
            
        return "\n".join(text_content)[:3000]  # Limit to 3000 chars for GPT
    
    def _extract_excel_content(self, excel_path):
        """Extract content from Excel for semantic analysis."""
        content = []
        
        try:
            wb = load_workbook(excel_path)
            ws = wb['Balance Sheet']
            
            for row in ws.iter_rows():
                row_content = []
                for cell in row:
                    if cell.value is not None:
                        row_content.append(str(cell.value))
                if row_content:
                    content.append(" | ".join(row_content))
                    
        except Exception as e:
            self.console.print(f"⚠️  Error extracting Excel content: {e}")
            
        return "\n".join(content)[:3000]  # Limit to 3000 chars for GPT
    
    def _create_validation_prompt(self, pdf_text, excel_content):
        """Create validation prompt for GPT-4o-mini."""
        return f"""
Please analyze the accuracy of financial data extraction from PDF to Excel format.

ORIGINAL PDF CONTENT:
{pdf_text}

EXTRACTED EXCEL CONTENT:
{excel_content}

Please evaluate:
1. Are the main financial sections (Assets, Liabilities, Equity, etc.) properly extracted?
2. Are account names preserved accurately?
3. Is the document structure and hierarchy maintained?
4. Are there any major omissions or misalignments?

Provide a validation score from 0-100% and explain your reasoning.
Format your response with the score at the end like: "SEMANTIC_SCORE: 85"
"""
    
    def _parse_gpt_score(self, gpt_response):
        """Extract numerical score from GPT response."""
        # Look for score pattern
        score_pattern = r'SEMANTIC_SCORE:\s*(\d+(?:\.\d+)?)'
        match = re.search(score_pattern, gpt_response)
        
        if match:
            return float(match.group(1))
        
        # Fallback: look for percentage patterns
        percentage_patterns = [
            r'(\d+(?:\.\d+)?)\s*%',
            r'score.*?(\d+(?:\.\d+)?)',
            r'(\d+(?:\.\d+)?)\s*out of 100'
        ]
        
        for pattern in percentage_patterns:
            match = re.search(pattern, gpt_response, re.IGNORECASE)
            if match:
                score = float(match.group(1))
                return min(score, 100)  # Cap at 100
                
        # Default fallback
        return 75.0
    
    def _calculate_hybrid_score(self, traditional_score, semantic_score):
        """Calculate final hybrid score with weighting."""
        # 70% traditional (numerical accuracy) + 30% semantic (structure/context)
        hybrid_score = (traditional_score * 0.7) + (semantic_score * 0.3)
        
        # Determine confidence level
        if hybrid_score >= 90:
            confidence = "Excellent"
        elif hybrid_score >= 80:
            confidence = "Good"
        elif hybrid_score >= 70:
            confidence = "Fair"
        else:
            confidence = "Poor"
            
        self.results['hybrid'] = {
            'final_score': round(hybrid_score, 1),
            'confidence_level': confidence,
            'traditional_weight': 70,
            'semantic_weight': 30,
            'traditional_score': traditional_score,
            'semantic_score': semantic_score
        }
        
        self.console.print(f"🎯 [bold green]Final Hybrid Score: {hybrid_score:.1f}% ({confidence})[/bold green]")
        return hybrid_score
    
    def _generate_report(self):
        """Generate comprehensive validation report."""
        # Create summary table
        table = Table(title="Hybrid Validation Summary", show_header=True)
        table.add_column("Validation Type", style="cyan", width=20)
        table.add_column("Score", style="green", width=10)
        table.add_column("Weight", style="yellow", width=10)
        table.add_column("Contribution", style="magenta", width=15)
        
        trad_score = self.results['traditional']['score']
        sem_score = self.results['semantic']['score']
        
        table.add_row("Traditional (Numerical)", f"{trad_score:.1f}%", "70%", f"{trad_score * 0.7:.1f}")
        table.add_row("Semantic (GPT-4o-mini)", f"{sem_score:.1f}%", "30%", f"{sem_score * 0.3:.1f}")
        table.add_row("[bold]Final Hybrid Score[/bold]", 
                     f"[bold green]{self.results['hybrid']['final_score']:.1f}%[/bold green]", 
                     "100%", 
                     f"[bold]{self.results['hybrid']['final_score']:.1f}[/bold]")
        
        self.console.print(table)
        
        # Display detailed findings
        if 'comparison' in self.results['traditional']:
            comp = self.results['traditional']['comparison']
            
            self.console.print(f"\n📊 [bold]Numerical Analysis Details:[/bold]")
            self.console.print(f"  • PDF Numbers Found: {comp['total_pdf_numbers']}")
            self.console.print(f"  • Excel Numbers Found: {comp['total_excel_numbers']}")
            self.console.print(f"  • Exact Matches: {comp['matched_numbers']}")
            
            if comp['missing_from_excel']:
                self.console.print(f"  • Missing from Excel: {len(comp['missing_from_excel'])} numbers")
                
        # Display GPT analysis
        if 'gpt_analysis' in self.results['semantic']:
            self.console.print(f"\n🤖 [bold]GPT-4o-mini Analysis:[/bold]")
            analysis_panel = Panel(self.results['semantic']['gpt_analysis'][:500] + "...", 
                                 title="Semantic Validation Details", 
                                 border_style="cyan")
            self.console.print(analysis_panel)

def main():
    """Main function to run hybrid validation."""
    import sys
    
    if len(sys.argv) != 3:
        print("Usage: python3 hybrid_validator.py <pdf_path> <excel_path>")
        print("\nExample:")
        print("  python3 hybrid_validator.py 'balance sheet.pdf' 'balance_sheet_extracted.xlsx'")
        return
    
    pdf_path = sys.argv[1]
    excel_path = sys.argv[2]
    
    # Check if files exist
    if not os.path.exists(pdf_path):
        print(f"❌ PDF file not found: {pdf_path}")
        return
        
    if not os.path.exists(excel_path):
        print(f"❌ Excel file not found: {excel_path}")
        return
    
    # Initialize validator and run
    validator = HybridValidator()
    results = validator.validate_extraction(pdf_path, excel_path)
    
    # Save results to JSON
    with open("validation_results.json", "w") as f:
        json.dump(results, f, indent=2)
    
    print(f"\n💾 Detailed results saved to: validation_results.json")

if __name__ == "__main__":
    main()