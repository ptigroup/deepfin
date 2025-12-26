#!/usr/bin/env python3
"""
Enterprise Output Manager

Professional-grade output management system with:
- Atomic operations (PROCESSING → SUCCESS/FAILED)
- Enhanced timestamp format (yyyymmdd_hhmmss)
- Run metadata and manifest generation
- Smart retention policy
- Status tracking and progress logging
- Multi-mode support (production/development/audit)
"""

import os
import json
import hashlib
import shutil
import time
import glob
from datetime import datetime, timedelta
from pathlib import Path
from typing import Dict, List, Optional, Any
from enum import Enum
from core import PATHS
from core.pipeline_logger import logger
import pandas as pd
import openpyxl
from openpyxl.utils.dataframe import dataframe_to_rows

class RunStatus(Enum):
    PROCESSING = "PROCESSING"
    SUCCESS = "SUCCESS"
    FAILED = "FAILED"

class OutputMode(Enum):
    PRODUCTION = "production"
    DEVELOPMENT = "development"
    AUDIT = "audit"

def safe_file_operation(operation_func, source, destination, operation_name="file operation"):
    """
    Safely perform file operations with same-file checking.
    
    Args:
        operation_func: The function to call (shutil.copy2, shutil.move, etc.)
        source: Source path
        destination: Destination path
        operation_name: Name of operation for error messages
    
    Returns:
        bool: True if operation was successful, False if skipped or failed
    """
    try:
        source_path = Path(source).resolve()
        dest_path = Path(destination).resolve()
        
        # Check if source and destination are the same file
        if source_path == dest_path:
            logger.debug(f"Skipping {operation_name}: source and destination are the same file")
            return True  # Not an error, just skip
        
        # Check if source exists (for copy/move operations)
        if not source_path.exists():
            logger.debug(f"Skipping {operation_name}: source file does not exist: {source_path}")
            return False
        
        # Perform the operation
        operation_func(str(source_path), str(dest_path))
        return True
        
    except Exception as e:
        error_msg = str(e)
        if "same file" in error_msg.lower():
            logger.debug(f"Skipping {operation_name}: attempting to {operation_name} file to itself")
            return True  # Not a fatal error
        else:
            print(f"  ❌ Error in {operation_name}: {e}")
            return False

class EnterpriseOutputManager:
    """Enterprise-grade output management system."""
    
    def __init__(self, base_output_dir: str = "output", mode: str = "production"):
        """Initialize the enterprise output manager."""
        self.base_output_dir = Path(base_output_dir)
        self.mode = OutputMode(mode.lower())
        
        # Configuration based on mode
        self.config = self._get_mode_config()
        
        # Current run information
        self.current_run_id = None
        self.current_run_dir = None
        self.run_start_time = None
        self.run_metadata = {}
        
        # Ensure base directories exist
        self._ensure_directories()
        
        logger.debug(f"🏢 Enterprise Output Manager initialized (mode: {self.mode.value})")
    
    def _get_mode_config(self) -> Dict[str, Any]:
        """Get configuration based on mode."""
        configs = {
            OutputMode.PRODUCTION: {
                "cleanup": True,
                "retention_policy": {
                    "keep_last_runs": 10,
                    "keep_daily_for_days": 30,
                    "keep_weekly_for_weeks": 12,
                    "compress_older_than_days": 7
                },
                "logging": "summary"
            },
            OutputMode.DEVELOPMENT: {
                "cleanup": False,
                "retention_policy": {
                    "keep_last_runs": 50,
                    "keep_daily_for_days": 90,
                    "keep_weekly_for_weeks": 52,
                    "compress_older_than_days": 30
                },
                "logging": "verbose"
            },
            OutputMode.AUDIT: {
                "cleanup": False,
                "retention_policy": {
                    "keep_last_runs": 100,
                    "keep_daily_for_days": 365,
                    "keep_weekly_for_weeks": 260,
                    "compress_older_than_days": 90
                },
                "logging": "full_trace"
            }
        }
        return configs[self.mode]
    
    def _ensure_directories(self):
        """Ensure all required directories exist."""
        directories = [
            self.base_output_dir,
            self.base_output_dir / "runs"
        ]
        
        # Audit directories are created within run directories as needed
        
        for directory in directories:
            directory.mkdir(parents=True, exist_ok=True)
    
    def _generate_timestamp(self) -> str:
        """Generate timestamp in yyyymmdd_hhmmss format."""
        return datetime.now().strftime("%Y%m%d_%H%M%S")
    
    def _generate_run_id(self) -> str:
        """Generate unique run ID with timestamp."""
        return self._generate_timestamp()
    
    def start_run(self, input_files: List[str], processing_mode: str = "multi_pdf") -> str:
        """
        Start a new processing run with atomic operation setup.
        
        Args:
            input_files: List of input file paths
            processing_mode: Type of processing (multi_pdf, single_pdf, etc.)
            
        Returns:
            run_id: Unique identifier for this run
        """
        self.current_run_id = self._generate_run_id()
        self.run_start_time = time.time()
        
        # Create PROCESSING directory (atomic operation)
        self.current_run_dir = self.base_output_dir / "runs" / f"{self.current_run_id}_{RunStatus.PROCESSING.value}"
        self.current_run_dir.mkdir(parents=True, exist_ok=True)
        
        # Initialize run metadata
        self.run_metadata = {
            "run_id": self.current_run_id,
            "start_time": datetime.now().isoformat(),
            "input_files": input_files,
            "processing_mode": processing_mode,
            "mode": self.mode.value,
            "status": RunStatus.PROCESSING.value,
            "pipeline_stages": [],
            "performance_metrics": {},
            "file_manifest": {}
        }
        
        # Save initial status
        self._save_run_status()
        self._update_latest_status()
        
        logger.debug(f"🚀 Started run {self.current_run_id} (mode: {processing_mode})")
        return self.current_run_id
    
    def log_stage_completion(self, stage_name: str, duration: float, output_files: List[str] = None):
        """Log completion of a pipeline stage."""
        stage_info = {
            "stage": stage_name,
            "completion_time": datetime.now().isoformat(),
            "duration_seconds": duration,
            "output_files": output_files or []
        }
        
        self.run_metadata["pipeline_stages"].append(stage_info)
        self.run_metadata["performance_metrics"][stage_name] = duration
        
        # Update status file
        self._save_run_status()
        
        if self.config["logging"] in ["verbose", "full_trace"]:
            print(f"  ✅ Stage '{stage_name}' completed in {duration:.1f}s")
    
    def register_file(self, file_path: str, file_type: str, stage: str, metadata: Dict = None):
        """Register a file in the run manifest."""
        file_path_obj = Path(file_path)
        
        if file_path_obj.exists():
            # Calculate checksum
            checksum = self._calculate_checksum(file_path)
            
            file_info = {
                "path": str(file_path),
                "type": file_type,
                "stage": stage,
                "size_bytes": file_path_obj.stat().st_size,
                "checksum": checksum,
                "created_time": datetime.now().isoformat(),
                "metadata": metadata or {}
            }
            
            if file_type not in self.run_metadata["file_manifest"]:
                self.run_metadata["file_manifest"][file_type] = []
            
            self.run_metadata["file_manifest"][file_type].append(file_info)
    
    def complete_run(self, final_files: Dict[str, str], success: bool = True) -> str:
        """
        Complete the run and organize files atomically.
        
        Args:
            final_files: Dictionary of final output files {"type": "path"}
            success: Whether the run was successful
            
        Returns:
            final_run_dir: Path to final run directory
        """
        run_duration = time.time() - self.run_start_time
        
        # Update final metadata
        self.run_metadata.update({
            "end_time": datetime.now().isoformat(),
            "duration_seconds": run_duration,
            "status": RunStatus.SUCCESS.value if success else RunStatus.FAILED.value,
            "final_files": final_files
        })
        
        # Determine final directory name
        final_status = RunStatus.SUCCESS if success else RunStatus.FAILED
        final_run_dir = self.base_output_dir / "runs" / f"{self.current_run_id}_{final_status.value}"
        
        try:
            # Step 1: Organize intermediate files based on mode
            self._organize_intermediate_files()
            
            # Step 2: Move/copy final files to run directory
            self._organize_run_files(final_files)
            
            # Step 3: Generate manifest and checksums
            self._generate_manifest()
            self._generate_checksums()
            
            # Step 4: Atomic move: PROCESSING → SUCCESS/FAILED
            if self.current_run_dir.exists():
                safe_file_operation(shutil.move, str(self.current_run_dir), str(final_run_dir), "atomic move to final state")
            
            # Step 5: Always update status files for user visibility
            self._generate_run_summary(self.base_output_dir)
            self._generate_comprehensive_status(final_files)
            
            # Step 6: Update LATEST files if successful
            if success:
                self._update_latest_files(final_files)
                self._cleanup_intermediate_files()
                self._cleanup_if_needed()
            
            # Step 6: Update status
            self._update_latest_status()
            
            logger.debug(f"✅ Run {self.current_run_id} completed successfully ({run_duration:.1f}s)")
            return str(final_run_dir)
            
        except Exception as e:
            logger.debug(f"❌ Error completing run {self.current_run_id}: {e}")
            # Move to FAILED directory
            failed_run_dir = self.base_output_dir / "runs" / f"{self.current_run_id}_FAILED"
            if self.current_run_dir.exists():
                safe_file_operation(shutil.move, str(self.current_run_dir), str(failed_run_dir), "move to FAILED state")
            return str(failed_run_dir)
    
    def _organize_run_files(self, final_files: Dict[str, str]):
        """Organize files in the run directory."""
        # Copy final files to run directory
        for file_type, file_path in final_files.items():
            if Path(file_path).exists():
                dest_path = self.current_run_dir / Path(file_path).name
                if safe_file_operation(shutil.copy2, file_path, dest_path, f"copy {file_type}"):
                    # Register in manifest only if copy was successful
                    self.register_file(str(dest_path), file_type, "final_output")
        
        # Audit files are already organized by _organize_intermediate_files()
    
    
    def _generate_manifest(self):
        """Generate complete run manifest."""
        manifest_path = self.current_run_dir / "manifest.json"
        with open(manifest_path, 'w', encoding='utf-8') as f:
            json.dump(self.run_metadata, f, indent=2)
    
    def _generate_checksums(self):
        """Generate MD5 checksums for all files in run directory."""
        checksums = {}
        
        for file_path in self.current_run_dir.rglob("*"):
            if file_path.is_file() and file_path.name != "checksums.md5":
                relative_path = file_path.relative_to(self.current_run_dir)
                checksums[str(relative_path)] = self._calculate_checksum(file_path)
        
        # Write checksums file
        checksum_path = self.current_run_dir / "checksums.md5"
        with open(checksum_path, 'w', encoding='utf-8') as f:
            for file_path, checksum in sorted(checksums.items()):
                f.write(f"{checksum}  {file_path}\n")
    
    def _calculate_checksum(self, file_path: str) -> str:
        """Calculate MD5 checksum of a file."""
        hash_md5 = hashlib.md5()
        try:
            with open(file_path, "rb") as f:
                for chunk in iter(lambda: f.read(4096), b""):
                    hash_md5.update(chunk)
            return hash_md5.hexdigest()
        except Exception:
            return "error"
    
    def _save_run_status(self):
        """Save current run status."""
        if self.current_run_dir:
            status_path = self.current_run_dir / "status.json"
            with open(status_path, 'w', encoding='utf-8') as f:
                json.dump({
                    "run_id": self.current_run_id,
                    "status": self.run_metadata["status"],
                    "start_time": self.run_metadata["start_time"],
                    "current_time": datetime.now().isoformat(),
                    "stages_completed": len(self.run_metadata["pipeline_stages"])
                }, f, indent=2)
    
    def _update_latest_status(self):
        """Update global latest status (now handled by _generate_comprehensive_status)."""
        # No-op: STATUS.json is now generated by _generate_comprehensive_status()
        # which is called from _update_latest_files()
        pass
    
    def _create_master_consolidated_excel(self, final_files: Dict[str, str], latest_dir: Path) -> Optional[str]:
        """Create a master Excel workbook combining all financial statements."""
        try:
            master_filename = f"Consolidated_Financial_Statements_{self.current_run_id}.xlsx"
            master_path = latest_dir / master_filename
            
            # Create a new workbook
            workbook = openpyxl.Workbook()
            
            # Remove default sheet
            workbook.remove(workbook.active)
            
            # Track which years are covered
            all_years = set()
            
            # Process each statement type
            for file_type, file_path in final_files.items():
                if not Path(file_path).exists():
                    continue
                
                try:
                    # Read the Excel file
                    df = pd.read_excel(file_path, engine='openpyxl')
                    
                    # Extract years from column headers
                    year_columns = [col for col in df.columns if str(col).isdigit() and len(str(col)) == 4]
                    all_years.update(year_columns)
                    
                    # Create worksheet name (capitalize and clean)
                    sheet_name = file_type.replace('_', ' ').title()
                    if len(sheet_name) > 31:  # Excel worksheet name limit
                        sheet_name = sheet_name[:31]
                    
                    # Create worksheet
                    worksheet = workbook.create_sheet(title=sheet_name)
                    
                    # Write dataframe to worksheet
                    for r in dataframe_to_rows(df, index=False, header=True):
                        worksheet.append(r)
                    
                    # Auto-adjust column widths
                    for column in worksheet.columns:
                        max_length = 0
                        column_letter = column[0].column_letter
                        
                        for cell in column:
                            try:
                                if len(str(cell.value)) > max_length:
                                    max_length = len(str(cell.value))
                            except:
                                pass
                        
                        adjusted_width = min(max_length + 2, 50)  # Cap at 50 characters
                        worksheet.column_dimensions[column_letter].width = adjusted_width
                    
                    print(f"  ✅ Added {sheet_name} worksheet to master file")
                    
                except Exception as e:
                    print(f"  ⚠️ Could not add {file_type} to master file: {e}")
                    continue
            
            # Create summary worksheet
            if workbook.worksheets:
                summary_sheet = workbook.create_sheet(title="Summary", index=0)
                
                # Add summary information
                summary_data = [
                    ["Consolidated Financial Statements"],
                    [""],
                    ["Run Information:"],
                    [f"Run ID: {self.current_run_id}"],
                    [f"Generated: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"],
                    [f"Processing Mode: {self.mode.value}"],
                    [""],
                    ["Covered Years:"],
                    [f"Years: {', '.join(sorted(str(year) for year in all_years))}"],
                    [""],
                    ["Worksheets Included:"],
                ]
                
                for sheet in workbook.worksheets[1:]:  # Skip summary sheet itself
                    summary_data.append([f"• {sheet.title}"])
                
                # Write summary to worksheet
                for row_data in summary_data:
                    summary_sheet.append(row_data)
                
                # Format summary sheet
                summary_sheet.column_dimensions['A'].width = 40
                for row in summary_sheet.iter_rows(min_row=1, max_row=1):
                    for cell in row:
                        cell.font = openpyxl.styles.Font(bold=True, size=14)
            
            # Save the master workbook
            workbook.save(master_path)
            print(f"📊 Created master consolidated file: {master_filename}")
            
            return str(master_path)
            
        except Exception as e:
            logger.debug(f"❌ Error creating master consolidated Excel: {e}")
            return None

    def _update_latest_files(self, final_files: Dict[str, str]):
        """Update main output directory with current results."""
        # Don't create master file in root output directory - it should stay in run directory
        # Master file is already created within the run directory by the pipeline
        
        # Summary and status files are now generated separately in complete_run
    
    def _generate_run_summary(self, latest_dir: Path):
        """Generate human-readable run summary."""
        summary_lines = [
            f"Financial Document Processing Summary",
            f"=" * 50,
            f"Run ID: {self.current_run_id}",
            f"Completed: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
            f"Mode: {self.mode.value}",
            f"Duration: {self.run_metadata.get('duration_seconds', 0):.1f} seconds",
            f"",
            f"Generated Files:",
        ]
        
        for file_path in sorted(latest_dir.glob("*.xlsx")):
            file_size = file_path.stat().st_size / 1024  # KB
            summary_lines.append(f"  ✅ {file_path.name} ({file_size:.1f} KB)")
        
        summary_lines.extend([
            f"",
            f"Pipeline Stages:",
        ])
        
        for stage in self.run_metadata.get("pipeline_stages", []):
            duration = stage.get("duration_seconds", 0)
            summary_lines.append(f"  • {stage['stage']}: {duration:.1f}s")
        
        # Write summary
        summary_path = latest_dir / "RUN_SUMMARY.txt"
        with open(summary_path, 'w', encoding='utf-8') as f:
            f.write('\n'.join(summary_lines))
        
        # Write metadata
        metadata_path = latest_dir / "METADATA.json"
        with open(metadata_path, 'w', encoding='utf-8') as f:
            json.dump(self.run_metadata, f, indent=2)
        
        # Note: RUN_SUMMARY.txt is now only stored in run-specific directories to avoid redundancy
        # Global summary information is available in STATUS.json
    
    def _generate_comprehensive_status(self, final_files: Dict[str, str]):
        """Generate comprehensive STATUS.json combining all status information."""
        # Create comprehensive status with simple access fields at top level
        comprehensive_status = {
            # Simple access fields (compatible with old LATEST_STATUS.json)
            "latest_run_id": self.current_run_id,
            "status": self.run_metadata.get("status", "UNKNOWN"),
            "timestamp": datetime.now().isoformat(),
            "mode": self.mode.value,
            
            # Comprehensive information from metadata
            "run_id": self.current_run_id,
            "start_time": self.run_metadata.get("start_time"),
            "end_time": self.run_metadata.get("end_time"),
            "duration_seconds": self.run_metadata.get("duration_seconds", 0),
            "input_files": self.run_metadata.get("input_files", []),
            "processing_mode": self.run_metadata.get("processing_mode"),
            "pipeline_stages": self.run_metadata.get("pipeline_stages", []),
            "performance_metrics": self.run_metadata.get("performance_metrics", {}),
            "file_manifest": self.run_metadata.get("file_manifest", {}),
            "final_files": final_files
        }
        
        # Write comprehensive status
        status_path = self.base_output_dir / "STATUS.json"
        with open(status_path, 'w', encoding='utf-8') as f:
            json.dump(comprehensive_status, f, indent=2)
        
        logger.debug(f"📊 Updated STATUS.json with comprehensive run information")
    
    def _organize_intermediate_files(self):
        """Organize intermediate files based on current mode."""
        if self.mode == OutputMode.AUDIT:
            # In audit mode, move all intermediate files to audit directory
            self._move_files_to_audit()
        # In production/development mode, we'll clean up later
    
    def _move_files_to_audit(self):
        """Move all intermediate files to run directory for audit mode."""
        if self.mode != OutputMode.AUDIT:
            return
        
        # In audit mode, just move files directly to the run directory without subfolders
        # The run directory itself serves as the audit container
        timestamp_patterns = [
            self.current_run_id.replace('_', ''),  # YYYYMMDDHHMMSS format
            self.current_run_id  # YYYYMMDD_HHMMSS format
        ]
        
        moved_files = 0
        for pattern in timestamp_patterns:
            for file_path in self.base_output_dir.glob(f"*{pattern}*"):
                if file_path.is_file() and not file_path.name.startswith('.'):
                    dest_path = self.current_run_dir / file_path.name
                    if safe_file_operation(shutil.move, str(file_path), str(dest_path), f"move audit file {file_path.name}"):
                        moved_files += 1
        
        if moved_files > 0:
            print(f"🗂️ Audit mode: Moved {moved_files} intermediate files to run directory")
    
    
    
    def _cleanup_intermediate_files(self):
        """Clean up intermediate files based on mode after successful run."""
        if self.mode == OutputMode.PRODUCTION:
            # In production mode, delete all intermediate files
            self._delete_intermediate_files()
        elif self.mode == OutputMode.DEVELOPMENT:
            # In development mode, keep some files for debugging
            self._cleanup_old_development_files()
        # Audit mode files are already moved to audit directory
    
    def _delete_intermediate_files(self):
        """Delete all intermediate files for production mode."""
        if self.mode != OutputMode.PRODUCTION:
            return
        
        # Find all files with current run timestamp in main output directory
        timestamp_patterns = [
            self.current_run_id.replace('_', ''),  # YYYYMMDDHHMMSS format
            self.current_run_id  # YYYYMMDD_HHMMSS format
        ]
        
        deleted_files = 0
        for pattern in timestamp_patterns:
            for file_path in self.base_output_dir.glob(f"*{pattern}*"):
                if file_path.is_file() and not file_path.name.startswith('.'):
                    # Don't delete final consolidated files
                    if 'multi-pdf-consolidated' in file_path.name and file_path.suffix == '.xlsx':
                        continue
                    try:
                        file_path.unlink()
                        deleted_files += 1
                    except Exception as e:
                        print(f"⚠️ Could not delete {file_path.name}: {e}")
        
        if deleted_files > 0:
            logger.debug(f"🧹 Production mode: Cleaned up {deleted_files} intermediate files")
    
    def _cleanup_old_development_files(self):
        """Clean up old development files while keeping recent ones."""
        if self.mode != OutputMode.DEVELOPMENT:
            return
        
        # Keep files from last 5 runs in development mode
        cutoff_time = datetime.now() - timedelta(hours=24)  # Keep last 24 hours
        
        deleted_files = 0
        for file_path in self.base_output_dir.glob("*"):
            if file_path.is_file() and not file_path.name.startswith('.'):
                # Skip final consolidated files and recent files
                if 'multi-pdf-consolidated' in file_path.name and file_path.suffix == '.xlsx':
                    continue
                
                file_time = datetime.fromtimestamp(file_path.stat().st_mtime)
                if file_time < cutoff_time:
                    try:
                        file_path.unlink()
                        deleted_files += 1
                    except Exception as e:
                        print(f"⚠️ Could not delete old file {file_path.name}: {e}")
        
        if deleted_files > 0:
            print(f"🧹 Development mode: Cleaned up {deleted_files} old files")
    
    def cleanup_existing_files(self):
        """Clean up all existing files in main output directory and organize by timestamp."""
        print(f"🧹 Cleaning up existing output directory...")
        
        # Get all files in main output directory (excluding subdirectories)
        all_files = [f for f in self.base_output_dir.iterdir() if f.is_file() and not f.name.startswith('.')]
        
        if not all_files:
            print("  ✅ Output directory already clean")
            return
        
        print(f"  📁 Found {len(all_files)} files to organize")
        
        # Group files by extracted timestamp
        files_by_timestamp = {}
        orphaned_files = []
        
        for file_path in all_files:
            timestamp = self._extract_timestamp_from_filename(file_path.name)
            if timestamp:
                if timestamp not in files_by_timestamp:
                    files_by_timestamp[timestamp] = []
                files_by_timestamp[timestamp].append(file_path)
            else:
                orphaned_files.append(file_path)
        
        # Organize files by timestamp into appropriate run directories or audit
        for timestamp, files in files_by_timestamp.items():
            self._organize_files_for_timestamp(timestamp, files)
        
        # Handle orphaned files
        if orphaned_files:
            self._handle_orphaned_files(orphaned_files)
        
        print(f"  ✅ Organized {len(all_files)} files")
    
    def _extract_timestamp_from_filename(self, filename: str) -> Optional[str]:
        """Extract timestamp from filename in DDMMYYYYHHMMSS format."""
        import re
        # Look for pattern like 21092025163040 (DDMMYYYYHHMMSS)
        match = re.search(r'(\d{14})', filename)
        if match:
            return match.group(1)
        return None
    
    def _organize_files_for_timestamp(self, timestamp: str, files: List[Path]):
        """Organize files for a specific timestamp into run directory or audit."""
        # Convert timestamp to run ID format (YYYYMMDD_HHMMSS)
        if len(timestamp) == 14:  # DDMMYYYYHHMMSS
            dd, mm, yyyy, hhmmss = timestamp[:2], timestamp[2:4], timestamp[4:8], timestamp[8:]
            run_id = f"{yyyy}{mm}{dd}_{hhmmss[:2]}{hhmmss[2:4]}{hhmmss[4:]}"
        else:
            return  # Invalid timestamp format
        
        # Find existing run directory for this timestamp
        runs_dir = self.base_output_dir / "runs"
        matching_run_dirs = list(runs_dir.glob(f"{run_id}_*"))
        
        if matching_run_dirs:
            # Use existing run directory
            target_run_dir = matching_run_dirs[0]
        else:
            # Create new SUCCESS run directory for historical files
            target_run_dir = runs_dir / f"{run_id}_SUCCESS"
            target_run_dir.mkdir(parents=True, exist_ok=True)
        
        # Move files based on mode
        if self.mode == OutputMode.AUDIT:
            # Move to audit directory
            audit_dir = self.base_output_dir / "audit" / run_id
            self._move_files_to_audit_for_timestamp(files, audit_dir)
        else:
            # Move appropriate files to run directory
            self._move_files_to_run_directory(files, target_run_dir)
    
    def _move_files_to_audit_for_timestamp(self, files: List[Path], audit_dir: Path):
        """Move files to audit directory for specific timestamp."""
        audit_dir.mkdir(parents=True, exist_ok=True)
        
        audit_subdirs = {
            "raw": audit_dir / "raw",
            "json_individual": audit_dir / "json_individual",
            "json_merged": audit_dir / "json_merged", 
            "json_consolidated": audit_dir / "json_consolidated",
            "excel_individual": audit_dir / "excel_individual"
        }
        
        for subdir in audit_subdirs.values():
            subdir.mkdir(exist_ok=True)
        
        for file_path in files:
            dest_dir = self._determine_audit_subdir(file_path.name, audit_subdirs)
            if dest_dir:
                dest_path = dest_dir / file_path.name
                safe_file_operation(shutil.move, str(file_path), str(dest_path), f"move audit file {file_path.name}")
    
    def _move_files_to_run_directory(self, files: List[Path], target_run_dir: Path):
        """Move only final consolidated files to run directory, delete others."""
        for file_path in files:
            if 'multi-pdf-consolidated' in file_path.name and file_path.suffix == '.xlsx':
                # Keep final consolidated Excel files
                dest_path = target_run_dir / file_path.name
                safe_file_operation(shutil.move, str(file_path), str(dest_path), f"move consolidated file {file_path.name}")
            else:
                # Delete intermediate files in production/development mode
                try:
                    file_path.unlink()
                except Exception as e:
                    print(f"⚠️ Could not delete {file_path.name}: {e}")
    
    def _handle_orphaned_files(self, orphaned_files: List[Path]):
        """Handle files that don't have recognizable timestamps by creating a FAILED run."""
        print(f"  ⚠️ Found {len(orphaned_files)} orphaned files without timestamps")
        
        # Create FAILED run directory for orphaned files
        failed_timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        failed_run_dir = self.base_output_dir / "runs" / f"{failed_timestamp}_FAILED"
        failed_run_dir.mkdir(parents=True, exist_ok=True)
        
        for file_path in orphaned_files:
            try:
                dest_path = failed_run_dir / file_path.name
                shutil.move(str(file_path), str(dest_path))
            except Exception as e:
                print(f"⚠️ Could not move orphaned file {file_path.name}: {e}")
        
        print(f"  📁 Moved orphaned files to FAILED run: {failed_run_dir.name}")

    def _cleanup_if_needed(self):
        """Apply retention policy and cleanup old runs."""
        if not self.config["cleanup"]:
            return
        
        logger.debug("🧹 Applying retention policy...")
        
        # Get all run directories
        runs_dir = self.base_output_dir / "runs"
        run_dirs = [d for d in runs_dir.iterdir() if d.is_dir()]
        
        # Parse run directories
        run_info = []
        for run_dir in run_dirs:
            try:
                name_parts = run_dir.name.split('_')
                if len(name_parts) >= 3:
                    timestamp_str = f"{name_parts[0]}_{name_parts[1]}"
                    status = name_parts[2]
                    timestamp = datetime.strptime(timestamp_str, "%Y%m%d_%H%M%S")
                    run_info.append({
                        "dir": run_dir,
                        "timestamp": timestamp,
                        "status": status
                    })
            except (ValueError, IndexError):
                continue
        
        # Sort by timestamp (newest first)
        run_info.sort(key=lambda x: x["timestamp"], reverse=True)
        
        # Apply retention policy
        retention = self.config["retention_policy"]
        now = datetime.now()
        
        for i, run in enumerate(run_info):
            age_days = (now - run["timestamp"]).days
            should_keep = False
            
            # Keep last N runs
            if i < retention["keep_last_runs"]:
                should_keep = True
            # Keep daily for X days
            elif age_days < retention["keep_daily_for_days"]:
                should_keep = True
            # Keep weekly for X weeks  
            elif age_days < retention["keep_weekly_for_weeks"] * 7:
                # Check if this is the only run for this week
                week_start = run["timestamp"] - timedelta(days=run["timestamp"].weekday())
                week_runs = [r for r in run_info if week_start <= r["timestamp"] < week_start + timedelta(days=7)]
                if run == min(week_runs, key=lambda x: x["timestamp"]):
                    should_keep = True
            
            if not should_keep:
                print(f"  🗑️ Removing old run: {run['dir'].name}")
                shutil.rmtree(run["dir"])
            elif age_days > retention["compress_older_than_days"]:
                # TODO: Implement compression for old runs
                pass

def main():
    """Test the enterprise output manager."""
    manager = EnterpriseOutputManager(mode="development")
    
    # Test run
    run_id = manager.start_run(["test1.pdf", "test2.pdf"], "multi_pdf")
    
    # Simulate stages
    manager.log_stage_completion("detection", 5.2, ["page39.txt", "page40.txt"])
    manager.log_stage_completion("extraction", 15.8, ["data.json"])
    manager.log_stage_completion("consolidation", 2.1, ["final.xlsx"])
    
    # Complete run
    final_files = {
        "income_statement": "output/income_statement.xlsx",
        "balance_sheet": "output/balance_sheet.xlsx"
    }
    
    manager.complete_run(final_files, success=True)

if __name__ == "__main__":
    main()